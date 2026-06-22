"""
Terra Roxa System - Sistema Completo de Gestão de Fazenda Leiteira
"""
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_migrate import Migrate
from datetime import datetime, date, timedelta
from io import BytesIO
import calendar
import os
from werkzeug.middleware.proxy_fix import ProxyFix

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)
app.secret_key = os.environ.get('SECRET_KEY', 'terra-roxa-sistema-2024')

# Configuração de banco: PostgreSQL (produção) ou SQLite (desenvolvimento)
database_url = os.environ.get('DATABASE_URL')
if database_url:
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///gadoleiteiro.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configurações de produção
if os.environ.get('RENDER') or os.environ.get('FLASK_ENV') == 'production':
    app.config['SESSION_COOKIE_SECURE'] = True
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['JSON_AS_ASCII'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ========== MODELOS ==========

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    nome = db.Column(db.String(150))
    senha_hash = db.Column(db.String(500))
    is_admin = db.Column(db.Boolean, default=False)
    role = db.Column(db.String(50), default='operador')  # admin, gerente, operador, visualizador
    ativo = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    acao = db.Column(db.String(200))
    detalhes = db.Column(db.Text)
    ip_address = db.Column(db.String(50))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship('User', backref='logs')

class Animal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    brinco = db.Column(db.String(50), unique=True, nullable=False)
    raca = db.Column(db.String(100))
    sexo = db.Column(db.String(10))  # macho, femea
    lote = db.Column(db.String(50))  # filhote
    ativo = db.Column(db.Boolean, default=True)
    # Reprodução
    data_ultima_inseminacao = db.Column(db.Date)
    data_parto_prevista = db.Column(db.Date)
    status_reproducao = db.Column(db.String(50), default='vazio')  # vazio, prenha, seca, etc.
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class SaudeAnimal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    animal_id = db.Column(db.Integer, db.ForeignKey('animal.id'), nullable=False)
    tipo = db.Column(db.String(100))  # vacinação, vermifugação, doença, etc.
    descricao = db.Column(db.Text)
    data_aplicacao = db.Column(db.Date, nullable=False)
    proxima_dose = db.Column(db.Date)
    custo = db.Column(db.Numeric(10, 2))
    observacoes = db.Column(db.Text)
    animal = db.relationship('Animal', backref='saude_registros')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TipoRacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    preco_kg = db.Column(db.Numeric(10, 2), nullable=False)
    tipo = db.Column(db.String(50), default='concentrado')  # concentrado, volumoso, suplementar
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ProducaoLeite(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    animal_id = db.Column(db.Integer, db.ForeignKey('animal.id'), nullable=True)
    litros = db.Column(db.Numeric(10, 2), nullable=False)
    gordura = db.Column(db.Numeric(5, 2))  # % gordura
    proteina = db.Column(db.Numeric(5, 2))  # % proteína
    ccs = db.Column(db.Numeric(10, 2))  # Contagem de Células Somáticas
    preco_venda = db.Column(db.Numeric(10, 4))
    total_receber = db.Column(db.Numeric(10, 2))
    data = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    animal = db.relationship('Animal', backref='producoes')

class ConsumoRacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    animal_id = db.Column(db.Integer, db.ForeignKey('animal.id'), nullable=False)
    tipo_racao_id = db.Column(db.Integer, db.ForeignKey('tipo_racao.id'), nullable=False)
    quantidade_kg = db.Column(db.Numeric(10, 2), nullable=False)
    data = db.Column(db.Date, nullable=False)
    custo = db.Column(db.Numeric(10, 2), nullable=False)
    eficiencia = db.Column(db.Numeric(10, 4))  # litros/kg ração
    animal = db.relationship('Animal', backref='consumos')
    tipo_racao = db.relationship('TipoRacao', backref='consumos')

class PrecoLeite(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    preco_litro = db.Column(db.Numeric(10, 4), nullable=False)
    data_vigencia = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Despesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Numeric(10, 2), nullable=False)
    categoria = db.Column(db.String(50))  # racao, energia, pessoal, veterinario, etc.
    data = db.Column(db.Date, nullable=False)
    observacoes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Orcamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ano = db.Column(db.Integer, nullable=False)
    mes = db.Column(db.Integer, nullable=False)
    categoria = db.Column(db.String(50))
    valor_previsto = db.Column(db.Numeric(10, 2))
    valor_realizado = db.Column(db.Numeric(10, 2))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    telefone = db.Column(db.String(50))
    email = db.Column(db.String(150))
    endereco = db.Column(db.String(300))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    vendas = db.relationship('VendaAvulsa', backref='cliente', lazy=True)

class VendaAvulsa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    data = db.Column(db.Date, nullable=False)
    litros = db.Column(db.Numeric(10, 2), nullable=False)
    valor_litro = db.Column(db.Numeric(10, 4), nullable=False)
    total = db.Column(db.Numeric(10, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ========== FUNÇÕES AUXILIARES ==========

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.context_processor
def inject_cotacao():
    hoje = date.today()
    preco = get_preco_vigente(hoje)
    data_fmt = hoje.strftime('%d/%m/%Y')
    alertas = SaudeAnimal.query.filter(
        SaudeAnimal.proxima_dose <= hoje + timedelta(days=7),
        SaudeAnimal.proxima_dose >= hoje
    ).count()
    return dict(preco_vigente_global=preco, data_cotacao_global=data_fmt, qtd_alertas_saude=alertas, date=date)

def log_auditoria(acao, detalhes='', user_id=None):
    if user_id is None and current_user.is_authenticated:
        user_id = current_user.id
    log = AuditLog(
        user_id=user_id,
        acao=acao,
        detalhes=detalhes,
        ip_address=request.remote_addr
    )
    db.session.add(log)
    db.session.commit()

def get_preco_vigente(data_ref):
    preco = PrecoLeite.query.filter(PrecoLeite.data_vigencia <= data_ref).order_by(PrecoLeite.data_vigencia.desc()).first()
    return preco.preco_litro if preco else 0

def calcular_eficiencia_alimentar(animal_id, data_ini, data_fim):
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.animal_id == animal_id,
        ProducaoLeite.data >= data_ini,
        ProducaoLeite.data <= data_fim
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.animal_id == animal_id,
        ConsumoRacao.data >= data_ini,
        ConsumoRacao.data <= data_fim
    ).all()
    
    total_leite = sum(p.litros for p in producoes)
    total_racao = sum(c.quantidade_kg for c in consumos)
    
    return total_leite / total_racao if total_racao > 0 else 0

def calcular_custo_producao(data_ini, data_fim):
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini,
        ProducaoLeite.data <= data_fim
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini,
        ConsumoRacao.data <= data_fim
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini,
        Despesa.data <= data_fim
    ).all()
    
    total_leite = sum(p.litros for p in producoes)
    total_custos = sum(c.custo for c in consumos) + sum(d.valor for d in despesas)
    
    return total_custos / total_leite if total_leite > 0 else 0

def evolucao_custo_litro(dias=30):
    arrays_datas, arrays_valores = [], []
    for i in range(dias - 1, -1, -1):
        d = date.today() - timedelta(days=i)
        arrays_datas.append(d.strftime('%d/%m'))
        custo = calcular_custo_producao(d, d)
        arrays_valores.append(round(custo, 4))
    return arrays_datas, arrays_valores

def evolucao_custo_litro_periodo(data_ini, data_fim):
    arrays_datas, arrays_valores = [], []
    delta = (data_fim - data_ini).days
    if delta > 365:
        meses = sorted(set(
            [data_ini + timedelta(days=i) for i in range(delta + 1)]
        ), key=lambda x: x.strftime('%Y-%m'))
        meses_unicos = {}
        for i in range(delta + 1):
            d = data_ini + timedelta(days=i)
            mes_key = d.strftime('%Y-%m')
            if mes_key not in meses_unicos:
                meses_unicos[mes_key] = {'inicio': d}
        for mes_key in sorted(meses_unicos):
            m = meses_unicos[mes_key]
            primeiro = m['inicio']
            ultimo = (primeiro.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            if ultimo > data_fim:
                ultimo = data_fim
            arrays_datas.append(primeiro.strftime('%m/%Y'))
            custo = calcular_custo_producao(primeiro, ultimo)
            arrays_valores.append(round(custo, 4))
    else:
        for i in range(delta + 1):
            d = data_ini + timedelta(days=i)
            arrays_datas.append(d.strftime('%d/%m'))
            custo = calcular_custo_producao(d, d)
            arrays_valores.append(round(custo, 4))
    return arrays_datas, arrays_valores

# ========== ROTAS ==========

@app.route('/health')
def health():
    return {'status': 'ok'}, 200

@app.route('/api/clima')
def api_clima():
    import requests
    WMO_CODES = {
        0: 'ceu limpo', 1: 'principalmente limpo', 2: 'parcialmente nublado', 3: 'nublado',
        45: 'nevoeiro', 48: 'nevoeiro denso',
        51: 'chuvisco leve', 53: 'chuvisco moderado', 55: 'chuvisco intenso',
        61: 'chuva leve', 63: 'chuva moderada', 65: 'chuva forte',
        66: 'chuva congelante leve', 67: 'chuva congelante forte',
        71: 'neve leve', 73: 'neve moderada', 75: 'neve intensa',
        77: 'graos de neve',
        80: 'pancadas de chuva leves', 81: 'pancadas de chuva moderadas', 82: 'pancadas de chuva violentas',
        85: 'pancadas de neve leves', 86: 'pancadas de neve intensas',
        95: 'tempestade', 96: 'tempestade com granizo leve', 99: 'tempestade com granizo forte'
    }
    WMO_ICONS = {
        0: '01d', 1: '02d', 2: '03d', 3: '04d',
        45: '50d', 48: '50d',
        51: '09d', 53: '09d', 55: '09d',
        61: '10d', 63: '10d', 65: '10d',
        66: '13d', 67: '13d',
        71: '13d', 73: '13d', 75: '13d', 77: '13d',
        80: '09d', 81: '09d', 82: '09d',
        85: '13d', 86: '13d',
        95: '11d', 96: '11d', 99: '11d'
    }
    try:
        LAT, LON = '-7.9268788', '-61.571846'
        url = f'https://api.open-meteo.com/v1/forecast?latitude={LAT}&longitude={LON}&current=temperature_2m,relative_humidity_2m,apparent_temperature,weather_code'
        resp = requests.get(url, timeout=5)
        data = resp.json()['current']
        code = data['weather_code']
        return {
            'temp': data['temperature_2m'],
            'description': WMO_CODES.get(code, 'nublado'),
            'humidity': data['relative_humidity_2m'],
            'icon': WMO_ICONS.get(code, '04d'),
            'location': 'Santo Antônio do Matupi - AM',
            'feels_like': data['apparent_temperature']
        }
    except:
        return {'temp': 25, 'description': 'ensolarado', 'humidity': 60, 'icon': '01d', 'location': 'Santo Antônio do Matupi - AM', 'feels_like': 24}

@app.route('/api/atualizar-preco', methods=['POST'])
@login_required
def atualizar_preco():
    try:
        data = request.get_json()
        novo_preco = float(data.get('preco', 0))
        
        if novo_preco <= 0:
            return {'success': False, 'error': 'Preço inválido'}
        
        # Atualizar ou criar novo preço vigente
        hoje = date.today()
        preco_existente = PrecoLeite.query.filter(PrecoLeite.data_vigencia <= hoje).order_by(PrecoLeite.data_vigencia.desc()).first()
        
        if preco_existente:
            preco_existente.preco_litro = novo_preco
        else:
            novo = PrecoLeite(data_vigencia=hoje, preco_litro=novo_preco)
            db.session.add(novo)
        
        db.session.commit()
        log_auditoria('Atualizar Preço Leite', f'Novo preço: R$ {novo_preco}')
        
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/api/cotacao-leite')
def api_cotacao_leite():
    try:
        hoje = date.today()
        preco = get_preco_vigente(hoje)
        preco_obj = PrecoLeite.query.filter(PrecoLeite.data_vigencia <= hoje).order_by(PrecoLeite.data_vigencia.desc()).first()
        data_cotacao = preco_obj.data_vigencia.strftime('%d/%m/%Y') if preco_obj else hoje.strftime('%d/%m/%Y')
        
        # Simular tendência (em produção, usaria dados históricos)
        import random
        tendencia = 'alta' if random.random() > 0.5 else 'baixa'
        
        return {
            'preco': float(preco),
            'data': data_cotacao,
            'tendencia': tendencia
        }
    except Exception as e:
        return {'preco': 2.50, 'data': date.today().strftime('%d/%m/%Y'), 'tendencia': 'estavel'}

@app.route('/')
@login_required
def index():
    today = date.today()
    
    # Métricas de hoje
    producoesHoje = ProducaoLeite.query.filter_by(data=today).all()
    total_litros = sum(p.litros for p in producoesHoje)
    num_animais_produzindo = len(set(p.animal_id for p in producoesHoje if p.animal_id))
    
    consumoHoje = ConsumoRacao.query.filter_by(data=today).all()
    total_custo_racao = sum(c.custo for c in consumoHoje)
    
    receita = 0
    for p in producoesHoje:
        preco = float(p.preco_venda) if p.preco_venda else float(get_preco_vigente(p.data))
        receita += float(p.litros) * preco
    
    lucro = receita - float(total_custo_racao)
    media_por_animal = total_litros / num_animais_produzindo if num_animais_produzindo > 0 else 0
    
    # Lucro mensal
    primeiro_dia_mes = date(today.year, today.month, 1)
    producoes_mes = ProducaoLeite.query.filter(ProducaoLeite.data >= primeiro_dia_mes).all()
    receita_mensal = 0
    for p in producoes_mes:
        preco = float(p.preco_venda) if p.preco_venda else float(get_preco_vigente(p.data))
        receita_mensal += float(p.litros) * preco
    custo_mensal = sum(c.custo for c in ConsumoRacao.query.filter(ConsumoRacao.data >= primeiro_dia_mes).all())
    lucro_mensal = receita_mensal - float(custo_mensal)
    
    # Últimos 7 dias
    ultimos_dias = []
    valores_producao = []
    custos_dia = []
    for i in range(6, -1, -1):
        d = date.today() - timedelta(days=i)
        ultimos_dias.append(d.strftime('%d/%m'))
        prods = ProducaoLeite.query.filter_by(data=d).all()
        valores_producao.append(sum(p.litros for p in prods))
        
        cons_dia = ConsumoRacao.query.filter_by(data=d).all()
        custos_dia.append(sum(c.custo for c in cons_dia))
    
    # Comparação com semana anterior
    semana_passada = []
    for i in range(13, 6, -1):
        d = date.today() - timedelta(days=i)
        prods = ProducaoLeite.query.filter_by(data=d).all()
        semana_passada.append(sum(p.litros for p in prods))
    
    total_semana_atual = sum(valores_producao)
    total_semana_passada = sum(semana_passada)
    variacao_semanal = ((total_semana_atual - total_semana_passada) / total_semana_passada * 100) if total_semana_passada > 0 else 0
    
    # Top 3 animais (maior produção na semana)
    from collections import defaultdict
    prod_por_animal = defaultdict(float)
    for i in range(6, -1, -1):
        d = date.today() - timedelta(days=i)
        prods = ProducaoLeite.query.filter_by(data=d).all()
        for p in prods:
            nome_animal = p.animal.nome if p.animal else 'Produção Geral'
            prod_por_animal[nome_animal] += float(p.litros)
    
    top_animais = sorted(prod_por_animal.items(), key=lambda x: x[1], reverse=True)[:3]
    
    # Tipos de ração mais consumidos (semana)
    racao_consumo = defaultdict(float)
    for i in range(6, -1, -1):
        d = date.today() - timedelta(days=i)
        cons = ConsumoRacao.query.filter_by(data=d).all()
        for c in cons:
            racao_consumo[c.tipo_racao.nome] += float(c.quantidade_kg)
    
    # Alertas de saúde (próximas vacinações/vermifugações)
    alertas_saude = SaudeAnimal.query.filter(
        SaudeAnimal.proxima_dose <= date.today() + timedelta(days=7),
        SaudeAnimal.proxima_dose >= date.today()
    ).all()
    
    custo_producao_hoje = calcular_custo_producao(today, today)
    
    custo_litro_dias, custo_litro_valores = evolucao_custo_litro(30)
    custo_litro_media = round(sum(custo_litro_valores) / len(custo_litro_valores), 4) if custo_litro_valores else 0
    
    # Resumo mensal de produção
    primeiro_dia_mes = date(today.year, today.month, 1)
    producoes_mes = ProducaoLeite.query.filter(ProducaoLeite.data >= primeiro_dia_mes).all()
    total_litros_mes = sum(p.litros for p in producoes_mes)
    receita_mensal = sum(float(p.litros) * float(p.preco_venda if p.preco_venda else get_preco_vigente(p.data)) for p in producoes_mes)
    
    # Produção mensal (últimos 12 meses)
    meses_labels = []
    valores_mensais = []
    for i in range(11, -1, -1):
        mes = today.month - i
        ano = today.year
        while mes < 1:
            mes += 12
            ano -= 1
        while mes > 12:
            mes -= 12
            ano += 1
        primeiro = date(ano, mes, 1)
        if mes == 12:
            ultimo = date(ano + 1, 1, 1) - timedelta(days=1)
        else:
            ultimo = date(ano, mes + 1, 1) - timedelta(days=1)
        prods = ProducaoLeite.query.filter(ProducaoLeite.data >= primeiro, ProducaoLeite.data <= ultimo).all()
        meses_labels.append(date(ano, mes, 1).strftime('%m/%Y'))
        valores_mensais.append(sum(p.litros for p in prods))

    # Cotação atual do leite
    preco_vigente_hoje = get_preco_vigente(today)
    data_preco_vigente = PrecoLeite.query.filter(PrecoLeite.data_vigencia <= today).order_by(PrecoLeite.data_vigencia.desc()).first()
    data_cotacao = data_preco_vigente.data_vigencia.strftime('%d/%m/%Y') if data_preco_vigente else today.strftime('%d/%m/%Y')
    
    return render_template('index.html', 
                           total_litros=total_litros,
                           receita=receita,
                           custo=total_custo_racao,
                           lucro=lucro,
                           lucro_mensal=lucro_mensal,
                           media_por_animal=media_por_animal,
                           num_animais_produzindo=num_animais_produzindo,
                           dias=ultimos_dias,
                           valores_producao=valores_producao,
                           custos_dia=custos_dia,
                           variacao_semanal=variacao_semanal,
                           top_animais=top_animais,
                           racao_consumo=dict(racao_consumo),
                           alertas_saude=alertas_saude,
                           custo_producao=custo_producao_hoje,
                           total_animais=Animal.query.filter_by(ativo=True).count(),
                           preco_vigente=preco_vigente_hoje,
                           data_cotacao=data_cotacao,
                           total_litros_mes=total_litros_mes,
                           receita_mensal_resumo=receita_mensal,
                           receita_mensal=receita_mensal,
                           custo_litro_dias=custo_litro_dias,
                           custo_litro_valores=custo_litro_valores,
                            custo_litro_media=custo_litro_media,
                            meses_labels=meses_labels,
                            valores_mensais=valores_mensais)
 
@app.route('/relatorios/pdf')
@login_required
def relatorio_pdf():
    from datetime import datetime
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini_date, ConsumoRacao.data <= data_fim_date
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini_date, Despesa.data <= data_fim_date
    ).all()
    
    total_litros = sum(p.litros for p in producoes)
    custo_total = sum(c.custo for c in consumos) + sum(d.valor for d in despesas)
    receita = sum(float(p.litros) * float(p.preco_venda if p.preco_venda else get_preco_vigente(p.data)) for p in producoes)
    lucro = receita - float(custo_total)
    
    # Tentar usar reportlab para PDF profissional
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.graphics.charts.lineplots import LinePlot
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics import renderPDF
        import tempfile
        import os
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#27AE60'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Cabeçalho
        elements.append(Paragraph('🐄 Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Gestão - {data_ini} a {data_fim}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Linha decorativa
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))
        
        # Resumo Executivo
        elements.append(Paragraph('Resumo Executivo', styles['Heading2']))
        
        # Tabela de métricas
        data = [
            ['Métrica', 'Valor', 'Detalhes'],
            ['Total Litros', f'{total_litros:.2f} L', f'{total_litros/max(1, (data_fim_date-data_ini_date).days):.2f} L/dia'],
            ['Receita Total', f'R$ {receita:.2f}', f'Média: R$ {receita/max(1,total_litros):.4f}/L'],
            ['Custo Total', f'R$ {custo_total:.2f}', f'{(custo_total/receita*100) if receita > 0 else 0:.1f}% da receita'],
            ['Lucro Líquido', f'R$ {lucro:.2f}', f'Margem: {(lucro/receita*100) if receita > 0 else 0:.1f}%']
        ]
        
        t = Table(data, colWidths=[4*cm, 4*cm, 6*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9F7')])
        ]))
        elements.append(t)
        elements.append(Spacer(1, 1*cm))
        
        # Produção por Animal (Top 5)
        if producoes:
            from collections import defaultdict
            prod_animal = defaultdict(float)
            for p in producoes:
                nome_animal = p.animal.nome if p.animal else 'Produção Geral'
                prod_animal[nome_animal] += float(p.litros)
            
            top5 = sorted(prod_animal.items(), key=lambda x: x[1], reverse=True)[:5]
            
            elements.append(Paragraph('Top 5 Animais - Produção', styles['Heading2']))
            
            data_animais = [['Animal', 'Total (L)', '% do Total']]
            for animal, prod in top5:
                pct = (prod / total_litros * 100) if total_litros > 0 else 0
                data_animais.append([animal, f'{prod:.2f}', f'{pct:.1f}%'])
            
            t2 = Table(data_animais, colWidths=[6*cm, 4*cm, 4*cm])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF5FB')])
            ]))
            elements.append(t2)
            elements.append(Spacer(1, 1*cm))
        
        # Custos por Categoria
        if despesas:
            from collections import defaultdict
            custos_cat = defaultdict(float)
            for d in despesas:
                custos_cat[d.categoria or 'Outros'] += float(d.valor)
            
            elements.append(Paragraph('Custos por Categoria', styles['Heading2']))
            
            data_custos = [['Categoria', 'Valor (R$)', '% do Total']]
            for cat, val in sorted(custos_cat.items(), key=lambda x: x[1], reverse=True):
                pct = (val / custo_total * 100) if custo_total > 0 else 0
                data_custos.append([cat, f'{val:.2f}', f'{pct:.1f}%'])
            
            t3 = Table(data_custos, colWidths=[6*cm, 4*cm, 4*cm])
            t3.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F39C12')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF9E7')])
            ]))
            elements.append(t3)
        
        # Rodapé
        elements.append(Spacer(1, 2*cm))
        elements.append(Paragraph(f'Relatório gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', 
                                           styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_terra_roxa_{data_ini}_{data_fim}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
        
    except Exception as e:
        # Fallback para HTML
        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>Relatório Terra Roxa</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 40px; }}
                h1 {{ color: #27AE60; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: center; }}
                th {{ background: #27AE60; color: white; }}
                .positive {{ color: green; }}
                .negative {{ color: red; }}
            </style>
        </head>
        <body>
            <h1>🐄 Terra Roxa System</h1>
            <h3>Relatório de Gestão</h3>
            <p><strong>Período:</strong> {data_ini} a {data_fim}</p>
            <hr>
            <h4>Resumo Executivo</h4>
            <table>
                <tr><th>Métrica</th><th>Valor</th></tr>
                <tr><td>Total Litros</td><td>{total_litros:.2f} L</td></tr>
                <tr><td>Receita Total</td><td>R$ {receita:.2f}</td></tr>
                <tr><td>Custo Total</td><td>R$ {custo_total:.2f}</td></tr>
                <tr><td>Lucro Líquido</td><td class="{'positive' if lucro >= 0 else 'negative'}">R$ {lucro:.2f}</td></tr>
            </table>
            <p style="margin-top: 40px; font-size: 12px; color: #666;">
                Relatório gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System
            </p>
        </body>
        </html>
        '''
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_terra_roxa_{data_ini}_{data_fim}.html'
        response.headers['Content-Type'] = 'text/html'
        return response

@app.route('/relatorios/excel')
@login_required
def relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from io import BytesIO
    from decimal import Decimal

    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')

    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')

    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()

    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini_date, ConsumoRacao.data <= data_fim_date
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini_date, Despesa.data <= data_fim_date
    ).all()

    total_litros = sum(p.litros for p in producoes)
    custo_total = float(sum(c.custo for c in consumos)) + float(sum(d.valor for d in despesas))
    receita = sum(float(p.litros) * float(p.preco_venda if p.preco_venda else get_preco_vigente(p.data)) for p in producoes)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Resumo ---
    ws = wb.active
    ws.title = 'Resumo'
    ws.cell(1, 1, 'Terra Roxa System - Relatorio de Gestao').font = Font(bold=True, size=14, color='27AE60')
    ws.merge_cells('A1:B1')
    ws.cell(2, 1, f'Periodo: {data_ini} a {data_fim}')
    ws.merge_cells('A2:B2')

    headers = ['Metrica', 'Valor']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    rows = [
        ('Total Litros', f'{total_litros:.2f} L'),
        ('Receita Total', f'R$ {receita:.2f}'),
        ('Custo Total', f'R$ {custo_total:.2f}'),
        ('Lucro Liquido', f'R$ {receita - custo_total:.2f}'),
    ]
    for i, (k, v) in enumerate(rows, 5):
        ws.cell(i, 1, k).border = thin_border
        ws.cell(i, 2, v).border = thin_border

    # --- Producao ---
    ws2 = wb.create_sheet('Producao')
    headers2 = ['Data', 'Animal', 'Litros', 'Preco/L', 'Valor']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, p in enumerate(producoes, 2):
        prec = p.preco_venda if p.preco_venda else get_preco_vigente(p.data)
        nome = p.animal.nome if p.animal else 'Geral'
        ws2.cell(i, 1, p.data.isoformat()).border = thin_border
        ws2.cell(i, 2, nome).border = thin_border
        ws2.cell(i, 3, float(p.litros)).border = thin_border
        ws2.cell(i, 4, float(prec)).border = thin_border
        ws2.cell(i, 5, round(float(p.litros) * float(prec), 2)).border = thin_border

    # --- Custos ---
    ws3 = wb.create_sheet('Custos')
    headers3 = ['Tipo', 'Data', 'Categoria/Descricao', 'Valor']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    i = 2
    for c in consumos:
        ws3.cell(i, 1, 'Racao').border = thin_border
        ws3.cell(i, 2, c.data.isoformat()).border = thin_border
        ws3.cell(i, 3, c.tipo_racao.nome).border = thin_border
        ws3.cell(i, 4, float(c.custo)).border = thin_border
        i += 1
    for d in despesas:
        ws3.cell(i, 1, 'Despesa').border = thin_border
        ws3.cell(i, 2, d.data.isoformat()).border = thin_border
        ws3.cell(i, 3, d.categoria or 'Geral').border = thin_border
        ws3.cell(i, 4, float(d.valor)).border = thin_border
        i += 1

    for ws_ in [ws, ws2, ws3]:
        for col in range(1, 6):
            letter = chr(64 + col)
            ws_.column_dimensions[letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_terra_roxa_{data_ini}_{data_fim}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/cooperativas')
@login_required
def cooperativas():
    # Integração com cooperativas (simulado por enquanto)
    # Em produção, usaria APIs reais de cooperativas
    cooperativas_info = [
        {
            'nome': 'Coopela',
            'contato': '(11) 1234-5678',
            'email': 'contato@coopela.com.br',
            'preco_leite': 3.60,
            'cargas_disponiveis': 5,
            'distancia_km': 45
        },
        {
            'nome': 'Laticínios Boa Vista',
            'contato': '(11) 9876-5432',
            'email': 'comercial@laticinios.com.br',
            'preco_leite': 3.55,
            'cargas_disponiveis': 3,
            'distancia_km': 60
        }
    ]
    return render_template('cooperativas.html', cooperativas=cooperativas_info)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = User.query.filter_by(email=email).first()
        if user and user.ativo:
            from werkzeug.security import check_password_hash
            if check_password_hash(user.senha_hash, password):
                login_user(user)
                log_auditoria('Login realizado', f'Usuário {email} fez login')
                return redirect(url_for('index'))
        
        flash('Email ou senha inválidos', 'danger')
        return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        email = request.form.get('email')
        nome = request.form.get('nome')
        senha = request.form.get('senha')
        
        if User.query.filter_by(email=email).first():
            flash('Email já cadastrado', 'danger')
            return redirect(url_for('register'))
        
        from werkzeug.security import generate_password_hash
        user = User(
            email=email,
            nome=nome,
            senha_hash=generate_password_hash(senha),
            is_admin=False,
            role='operador'
        )
        db.session.add(user)
        db.session.commit()
        log_auditoria('Usuário criado', f'Novo usuário: {email}')
        
        flash('Usuário criado com sucesso! Faça login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    log_auditoria('Logout realizado', f'Usuário {current_user.email} saiu')
    logout_user()
    return redirect(url_for('login'))

# ========== PRODUÇÃO ==========

@app.route('/producao', methods=['GET', 'POST'])
@login_required
def producao():
    if request.method == 'POST':
        litros = float(request.form.get('litros'))
        data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        preco_venda = float(request.form.get('preco_venda')) if request.form.get('preco_venda') else 2.20
        total_receber = round(litros * preco_venda, 2)
        
        nova_producao = ProducaoLeite(
            animal_id=None, litros=litros, data=data, 
            preco_venda=preco_venda,
            total_receber=total_receber
        )
        db.session.add(nova_producao)
        db.session.commit()
        log_auditoria('Produção registrada', f'{litros}L a R$ {preco_venda}/L')
        flash('Produção registrada com sucesso!', 'success')
        return redirect(url_for('producao'))
    
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    ultimo_dia_mes = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1])
    
    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    
    if filtro_data_ini:
        data_ini = datetime.strptime(filtro_data_ini, '%Y-%m-%d').date()
    else:
        data_ini = primeiro_dia_mes
        filtro_data_ini = data_ini.strftime('%Y-%m-%d')
    
    if filtro_data_fim:
        data_fim = datetime.strptime(filtro_data_fim, '%Y-%m-%d').date()
    else:
        data_fim = ultimo_dia_mes
        filtro_data_fim = data_fim.strftime('%Y-%m-%d')
    
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini,
        ProducaoLeite.data <= data_fim
    ).order_by(ProducaoLeite.data.desc()).all()
    
    today = hoje.strftime('%Y-%m-%d')
    
    total_litros_geral = sum(float(p.litros) for p in producoes)
    total_receber_geral = sum(float(p.total_receber) for p in producoes if p.total_receber)
    
    return render_template('producao.html', 
                           producoes=producoes, 
                           today=today,
                           preco_padrao=2.20,
                           total_litros_geral=total_litros_geral,
                           total_receber_geral=total_receber_geral,
                           data_ini=filtro_data_ini,
                           data_fim=filtro_data_fim)

@app.route('/producao/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producao(id):
    producao = ProducaoLeite.query.get_or_404(id)
    if request.method == 'POST':
        producao.data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        producao.litros = float(request.form.get('litros'))
        producao.preco_venda = float(request.form.get('preco_venda'))
        producao.total_receber = round(producao.litros * producao.preco_venda, 2)
        db.session.commit()
        log_auditoria('Produção editada', f'{producao.litros}L a R$ {producao.preco_venda}/L')
        flash('Produção atualizada!', 'success')
        return redirect(url_for('producao'))
    return render_template('editar_producao.html', producao=producao)

@app.route('/producao/excluir/<int:id>')
@login_required
def excluir_producao(id):
    if current_user.role not in ['admin', 'gerente']:
        flash('Acesso restrito', 'danger')
        return redirect(url_for('producao'))
    producao = ProducaoLeite.query.get(id)
    db.session.delete(producao)
    db.session.commit()
    log_auditoria('Produção excluída', f'ID {id}')
    flash('Produção excluída!', 'success')
    return redirect(url_for('producao'))

# ========== EXPORTAÇÃO PRODUÇÃO ==========

@app.route('/producao/exportar/pdf')
@login_required
def producao_exportar_pdf():
    from datetime import datetime
    from decimal import Decimal
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).order_by(ProducaoLeite.data.asc()).all()
    total_litros = sum(p.litros for p in producoes)
    total_receber = sum(p.total_receber or 0 for p in producoes)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=20, textColor=colors.HexColor('#27AE60'),
                                     spaceAfter=20, alignment=TA_CENTER)
        elements.append(Paragraph('Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Produção - {data_ini} a {data_fim}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(f'<b>Total de Registros:</b> {len(producoes)}  |  <b>Total Litros:</b> {float(total_litros):.0f} L  |  <b>Total a Receber:</b> R$ {float(total_receber):.2f}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        data_table = [['Data', 'Litros', 'Preço/L', 'Total']]
        for p in producoes:
            data_table.append([
                p.data.strftime('%d/%m/%Y'),
                f'{float(p.litros):.1f}',
                f'R$ {float(p.preco_venda or 0):.2f}',
                f'R$ {float(p.total_receber or 0):.2f}'
            ])
        t = Table(data_table, colWidths=[4*cm, 3*cm, 3*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9F7')])
        ]))
        elements.append(t)
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', styles['Normal']))
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=producao_{data_ini}_{data_fim}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        html = f'''
        <html><head><title>Produção</title><style>
        body{{font-family:Arial;padding:40px}} h1{{color:#27AE60}}
        table{{border-collapse:collapse;width:100%;margin:20px 0}}
        th,td{{border:1px solid #ddd;padding:12px;text-align:center}} th{{background:#27AE60;color:#fff}}
        </style></head><body>
        <h1>Terra Roxa System</h1><h3>Relatório de Produção</h3>
        <p><strong>Período:</strong> {data_ini} a {data_fim}</p>
        <p>Total: {len(producoes)} registros | {float(total_litros):.0f} L | R$ {float(total_receber):.2f}</p><hr>
        <table><tr><th>Data</th><th>Litros</th><th>Preço/L</th><th>Total</th></tr>'''
        for p in producoes:
            html += f'<tr><td>{p.data.strftime("%d/%m/%Y")}</td><td>{float(p.litros):.1f}</td><td>R$ {float(p.preco_venda or 0):.2f}</td><td>R$ {float(p.total_receber or 0):.2f}</td></tr>'
        html += '</table></body></html>'
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=producao_{data_ini}_{data_fim}.html'
        response.headers['Content-Type'] = 'text/html'
        return response

@app.route('/producao/exportar/excel')
@login_required
def producao_exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from decimal import Decimal
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).order_by(ProducaoLeite.data.asc()).all()
    total_litros = sum(p.litros for p in producoes)
    total_receber = sum(p.total_receber or 0 for p in producoes)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Producao'
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(1, 1, 'Terra Roxa System - Relatorio de Producao').font = Font(bold=True, size=14, color='27AE60')
    ws.merge_cells('A1:D1')
    ws.cell(2, 1, f'Periodo: {data_ini} a {data_fim}')
    ws.merge_cells('A2:D2')
    headers = ['Data', 'Litros', 'Preco/L', 'Total a Receber']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for i, p in enumerate(producoes, 5):
        ws.cell(i, 1, p.data.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(i, 2, float(p.litros)).border = thin_border
        ws.cell(i, 3, float(p.preco_venda or 0)).border = thin_border
        ws.cell(i, 4, float(p.total_receber or 0)).border = thin_border
    row = 5 + len(producoes)
    ws.cell(row, 1, 'TOTAIS').font = Font(bold=True)
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 2, float(total_litros)).font = Font(bold=True)
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 3).border = thin_border
    ws.cell(row, 4, float(total_receber)).font = Font(bold=True)
    ws.cell(row, 4).border = thin_border
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 18
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=producao_{data_ini}_{data_fim}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# ========== RAÇÃO ==========

@app.route('/racao', methods=['GET', 'POST'])
@login_required
def racao():
    if request.method == 'POST':
        nome = request.form.get('nome')
        preco_kg = float(request.form.get('preco_kg'))
        tipo = request.form.get('tipo', 'concentrado')
        
        novo_tipo = TipoRacao(nome=nome, preco_kg=preco_kg, tipo=tipo)
        db.session.add(novo_tipo)
        db.session.commit()
        log_auditoria('Tipo ração cadastrado', f'{nome} - R$ {preco_kg}/kg')
        flash('Tipo de ração cadastrado!', 'success')
        return redirect(url_for('racao'))
    
    tipos = TipoRacao.query.all()
    return render_template('racao.html', tipos=tipos)

@app.route('/racao/excluir/<int:id>')
@login_required
def excluir_racao(id):
    tipo = TipoRacao.query.get_or_404(id)
    db.session.delete(tipo)
    db.session.commit()
    log_auditoria('Tipo ração excluído', f'{tipo.nome}')
    flash('Tipo de ração excluído!', 'success')
    return redirect(url_for('racao'))

@app.route('/racao/consumo/excluir/<int:id>')
@login_required
def excluir_consumo_racao(id):
    consumo = ConsumoRacao.query.get_or_404(id)
    db.session.delete(consumo)
    db.session.commit()
    log_auditoria('Consumo ração excluído', f'ID {id}')
    flash('Consumo excluído!', 'success')
    return redirect(url_for('consumo_racao'))

@app.route('/racao/consumo', methods=['GET', 'POST'])
@login_required
def consumo_racao():
    if request.method == 'POST':
        animal_id = request.form.get('animal_id')
        tipo_racao_id = request.form.get('tipo_racao_id')
        quantidade_kg = float(request.form.get('quantidade_kg'))
        data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        
        tipo = TipoRacao.query.get(tipo_racao_id)
        custo = quantidade_kg * float(tipo.preco_kg)
        eficiencia = None
        if quantidade_kg > 0:
            producao_dia = ProducaoLeite.query.filter_by(animal_id=animal_id, data=data).first()
            if producao_dia:
                eficiencia = float(producao_dia.litros) / quantidade_kg
        
        novo_consumo = ConsumoRacao(
            animal_id=animal_id, tipo_racao_id=tipo_racao_id,
            quantidade_kg=quantidade_kg, data=data, custo=custo,
            eficiencia=eficiencia
        )
        db.session.add(novo_consumo)
        db.session.commit()
        log_auditoria('Consumo registrado', f'Animal ID {animal_id}, {quantidade_kg}kg')
        flash('Consumo registrado!', 'success')
        return redirect(url_for('consumo_racao'))
    
    animais = Animal.query.filter_by(ativo=True).all()
    tipos = TipoRacao.query.all()
    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    
    query = ConsumoRacao.query
    if filtro_data_ini:
        query = query.filter(ConsumoRacao.data >= datetime.strptime(filtro_data_ini, '%Y-%m-%d').date())
    if filtro_data_fim:
        query = query.filter(ConsumoRacao.data <= datetime.strptime(filtro_data_fim, '%Y-%m-%d').date())
    
    consumos = query.order_by(ConsumoRacao.data.desc()).all()
    
    today = date.today().strftime('%Y-%m-%d')
    return render_template('consumo_racao.html', animais=animais, tipos=tipos, consumos=consumos, today=today)

# ========== ANIMAIS ==========

@app.route('/animais', methods=['GET', 'POST'])
@login_required
def animais():
    if request.method == 'POST':
        nome = request.form.get('nome')
        brinco = request.form.get('brinco')
        if not brinco:
            ultimo = Animal.query.order_by(Animal.id.desc()).first()
            prox_id = (ultimo.id + 1) if ultimo else 1
            brinco = f'AUTO-{prox_id:04d}'
        raca = request.form.get('raca')
        lote = request.form.get('lote')
        sexo = request.form.get('sexo')
        novo_animal = Animal(
            nome=nome, brinco=brinco, raca=raca, lote=lote,
            sexo=sexo
        )
        db.session.add(novo_animal)
        db.session.commit()
        log_auditoria('Animal cadastrado', f'{nome} - Brinco {brinco}')
        flash('Animal cadastrado!', 'success')
        return redirect(url_for('animais'))
    
    animais = Animal.query.order_by(Animal.nome).all()
    return render_template('animais.html', animais=animais)

@app.route('/animais/editar/<int:id>', methods=['POST'])
@login_required
def editar_animal(id):
    animal = Animal.query.get_or_404(id)
    animal.nome = request.form.get('nome', animal.nome)
    animal.raca = request.form.get('raca', animal.raca)
    animal.sexo = request.form.get('sexo', animal.sexo)
    animal.lote = request.form.get('lote', animal.lote)
    animal.ativo = request.form.get('ativo') == 'on'
    db.session.commit()
    log_auditoria('Animal editado', f'{animal.nome} - Brinco {animal.brinco}')
    flash('Animal atualizado!', 'success')
    return redirect(url_for('animais'))

@app.route('/animais/excluir/<int:id>')
@login_required
def excluir_animal(id):
    animal = Animal.query.get_or_404(id)
    try:
        SaudeAnimal.query.filter_by(animal_id=id).delete()
        ConsumoRacao.query.filter_by(animal_id=id).delete()
        ProducaoLeite.query.filter_by(animal_id=id).delete()
        db.session.delete(animal)
        db.session.commit()
        log_auditoria('Animal excluído', f'{animal.nome} - Brinco {animal.brinco}')
        flash('Animal excluído!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir animal: {str(e)}', 'danger')
    return redirect(url_for('animais'))

@app.route('/animais/saude/<int:id>', methods=['GET', 'POST'])
@login_required
def saude_animal(id):
    animal = Animal.query.get_or_404(id)
    
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        descricao = request.form.get('descricao')
        data_aplicacao = datetime.strptime(request.form.get('data_aplicacao'), '%Y-%m-%d').date()
        proxima_dose = datetime.strptime(request.form.get('proxima_dose'), '%Y-%m-%d').date() if request.form.get('proxima_dose') else None
        custo = float(request.form.get('custo')) if request.form.get('custo') else None
        observacoes = request.form.get('observacoes')
        
        registro = SaudeAnimal(
            animal_id=id, tipo=tipo, descricao=descricao,
            data_aplicacao=data_aplicacao, proxima_dose=proxima_dose,
            custo=custo, observacoes=observacoes
        )
        db.session.add(registro)
        db.session.commit()
        log_auditoria('Registro saúde', f'Animal: {animal.nome}, Tipo: {tipo}')
        flash('Registro de saúde adicionado!', 'success')
        return redirect(url_for('saude_animal', id=id))
    
    registros = SaudeAnimal.query.filter_by(animal_id=id).order_by(SaudeAnimal.data_aplicacao.desc()).all()
    from datetime import date
    return render_template('saude_animal.html', animal=animal, registros=registros, today=date.today())

# ========== RELATÓRIOS ==========
@app.route('/relatorios')
@login_required
def relatorios():
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    tipo = request.args.get('tipo', 'producao')
    
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    if tipo == 'vendas-avulsas':
        vendas = VendaAvulsa.query.filter(
            VendaAvulsa.data >= data_ini_date, VendaAvulsa.data <= data_fim_date
        ).order_by(VendaAvulsa.data.desc()).all()
        
        total_litros_vendas = sum(float(v.litros) for v in vendas)
        total_valor_vendas = sum(float(v.total) for v in vendas)
        
        dias_grafico = []
        valores_venda = []
        delta = data_fim_date - data_ini_date
        if delta.days <= 31:
            for i in range(delta.days + 1):
                d = data_ini_date + timedelta(days=i)
                dias_grafico.append(d.strftime('%d/%m'))
                vals_dia = [float(v.total) for v in vendas if v.data == d]
                valores_venda.append(sum(vals_dia))
        
        vendas_por_mes = {}
        for v in vendas:
            mes_key = v.data.strftime('%Y-%m')
            vendas_por_mes[mes_key] = vendas_por_mes.get(mes_key, 0) + float(v.total)
        meses_vendas = sorted(vendas_por_mes.keys())
        if meses_vendas:
            meses_labels_vendas = [datetime.strptime(m + '-01', '%Y-%m-%d').strftime('%m/%Y') for m in meses_vendas]
            valores_vendas_mensais = [vendas_por_mes[m] for m in meses_vendas]
        else:
            meses_labels_vendas = []
            valores_vendas_mensais = []
        
        return render_template('relatorios.html',
                               data_ini=data_ini, data_fim=data_fim,
                               tipo=tipo,
                               vendas=vendas,
                               total_litros_vendas=total_litros_vendas,
                               total_valor_vendas=total_valor_vendas,
                               dias_grafico=dias_grafico,
                               valores_venda=valores_venda,
                               meses_labels_vendas=meses_labels_vendas,
                               valores_vendas_mensais=valores_vendas_mensais)
    
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini_date, ConsumoRacao.data <= data_fim_date
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini_date, Despesa.data <= data_fim_date
    ).all()
    
    total_litros = sum(p.litros for p in producoes)
    custo_total = sum(c.custo for c in consumos) + sum(d.valor for d in despesas)
    
    receita = 0
    for p in producoes:
        preco = float(p.preco_venda) if p.preco_venda else float(get_preco_vigente(p.data))
        receita += float(p.litros) * preco
    
    lucro = receita - float(custo_total)
    custo_producao = calcular_custo_producao(data_ini_date, data_fim_date)
    
    dias_grafico = []
    valores_prod = []
    custos_dia = []
    
    delta = data_fim_date - data_ini_date
    if delta.days <= 31:
        for i in range(delta.days + 1):
            d = data_ini_date + timedelta(days=i)
            dias_grafico.append(d.strftime('%d/%m'))
            
            prods_dia = [p.litros for p in producoes if p.data == d]
            valores_prod.append(sum(prods_dia))
            
            custos_dia_sum = [c.custo for c in consumos if c.data == d]
            custos_dia.append(sum(custos_dia_sum))
    
    custo_por_tipo = {}
    for c in consumos:
        tipo_nome = c.tipo_racao.nome
        custo_por_tipo[tipo_nome] = custo_por_tipo.get(tipo_nome, 0) + float(c.custo)

    producao_por_mes = {}
    for p in producoes:
        mes_key = p.data.strftime('%Y-%m')
        producao_por_mes[mes_key] = producao_por_mes.get(mes_key, 0) + float(p.litros)

    custo_por_mes = {}
    for c in consumos:
        mes_key = c.data.strftime('%Y-%m')
        custo_por_mes[mes_key] = custo_por_mes.get(mes_key, 0) + float(c.custo)
    for d in despesas:
        mes_key = d.data.strftime('%Y-%m')
        custo_por_mes[mes_key] = custo_por_mes.get(mes_key, 0) + float(d.valor)

    receita_por_mes = {}
    for p in producoes:
        mes_key = p.data.strftime('%Y-%m')
        preco = float(p.preco_venda) if p.preco_venda else float(get_preco_vigente(p.data))
        receita_por_mes[mes_key] = receita_por_mes.get(mes_key, 0) + float(p.litros) * preco

    todos_meses = sorted(set(list(producao_por_mes.keys()) + list(custo_por_mes.keys()) + list(receita_por_mes.keys())))
    if todos_meses:
        meses_labels = [datetime.strptime(m + '-01', '%Y-%m-%d').strftime('%m/%Y') for m in todos_meses]
        valores_mensais = [producao_por_mes.get(m, 0) for m in todos_meses]
        custo_mensal = [custo_por_mes.get(m, 0) for m in todos_meses]
        lucro_mensal = [receita_por_mes.get(m, 0) - custo_por_mes.get(m, 0) for m in todos_meses]
    else:
        meses_labels = []
        valores_mensais = []
        custo_mensal = []
        lucro_mensal = []
    
    preco_medio = get_preco_vigente(data_fim_date)

    rel_custo_litro_dias, rel_custo_litro_valores = evolucao_custo_litro_periodo(data_ini_date, data_fim_date)
    rel_custo_litro_media = round(sum(rel_custo_litro_valores) / len(rel_custo_litro_valores), 4) if rel_custo_litro_valores else 0

    return render_template('relatorios.html',
                           data_ini=data_ini, data_fim=data_fim,
                           tipo=tipo,
                           total_litros=total_litros, receita=receita,
                           custo=custo_total, lucro=lucro,
                           custo_producao=custo_producao,
                           dias_grafico=dias_grafico, valores_prod=valores_prod,
                           custos_dia=custos_dia, custo_por_tipo=custo_por_tipo,
                           preco_medio=preco_medio,
                           meses_labels=meses_labels, valores_mensais=valores_mensais,
                           custo_mensal=custo_mensal, lucro_mensal=lucro_mensal,
                           rel_custo_litro_dias=rel_custo_litro_dias,
                           rel_custo_litro_valores=rel_custo_litro_valores,
                           rel_custo_litro_media=rel_custo_litro_media)

# ========== FINANCEIRO ==========

@app.route('/financeiro/editar/<int:id>', methods=['POST'])
@login_required
def editar_despesa(id):
    despesa = Despesa.query.get_or_404(id)
    despesa.descricao = request.form.get('descricao', despesa.descricao)
    despesa.categoria = request.form.get('categoria', despesa.categoria)
    despesa.valor = float(request.form.get('valor', despesa.valor))
    despesa.data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
    despesa.observacoes = request.form.get('observacoes', despesa.observacoes)
    db.session.commit()
    log_auditoria('Despesa editada', f'{despesa.descricao} - R$ {despesa.valor}')
    flash('Despesa atualizada!', 'success')
    return redirect(url_for('financeiro'))

@app.route('/financeiro/excluir/<int:id>')
@login_required
def excluir_despesa(id):
    despesa = Despesa.query.get_or_404(id)
    db.session.delete(despesa)
    db.session.commit()
    log_auditoria('Despesa excluída', f'{despesa.descricao} - R$ {despesa.valor}')
    flash('Despesa excluída!', 'success')
    return redirect(url_for('financeiro'))

@app.route('/financeiro', methods=['GET', 'POST'])
@login_required
def financeiro():
    if request.method == 'POST':
        descricao = request.form.get('descricao')
        valor = float(request.form.get('valor'))
        categoria = request.form.get('categoria')
        data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        observacoes = request.form.get('observacoes')
        
        despesa = Despesa(
            descricao=descricao, valor=valor, categoria=categoria,
            data=data, observacoes=observacoes
        )
        db.session.add(despesa)
        db.session.commit()
        log_auditoria('Despesa registrada', f'{descricao} - R$ {valor}')
        flash('Despesa registrada!', 'success')
        return redirect(url_for('financeiro'))
    
    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    filtro_categoria = request.args.get('categoria')
    
    query = Despesa.query
    if filtro_data_ini:
        query = query.filter(Despesa.data >= datetime.strptime(filtro_data_ini, '%Y-%m-%d').date())
    if filtro_data_fim:
        query = query.filter(Despesa.data <= datetime.strptime(filtro_data_fim, '%Y-%m-%d').date())
    if filtro_categoria:
        query = query.filter(Despesa.categoria == filtro_categoria)
    
    despesas = query.order_by(Despesa.data.desc()).all()
    
    total_despesas = sum(d.valor for d in despesas)
    
    from datetime import date
    return render_template('financeiro.html', despesas=despesas, total_despesas=total_despesas, today=date.today())

@app.route('/ajustes/reset', methods=['POST'])
@login_required
def reset_dados():
    if current_user.role not in ['admin', 'gerente']:
        flash('Acesso restrito', 'danger')
        return redirect(url_for('ajustes'))

    tipo = request.form.get('tipo', '')
    try:
        if tipo == 'producao':
            ProducaoLeite.query.delete()
            flash('Dados de produção limpos!', 'success')
        elif tipo == 'racao':
            ConsumoRacao.query.delete()
            flash('Dados de ração limpos!', 'success')
        elif tipo == 'todos':
            ProducaoLeite.query.delete()
            ConsumoRacao.query.delete()
            Despesa.query.delete()
            VendaAvulsa.query.delete()
            SaudeAnimal.query.delete()
            flash('Todos os dados foram limpos!', 'success')
        db.session.commit()
        log_auditoria('Reset dados', f'Tipo: {tipo}')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao limpar dados: {str(e)}', 'danger')
    return redirect(url_for('ajustes'))

# ========== CONFIGURAÇÕES ==========

@app.route('/orcamento', methods=['GET', 'POST'])
@login_required
def orcamento():
    if request.method == 'POST':
        ano = int(request.form.get('ano'))
        mes = int(request.form.get('mes'))
        categoria = request.form.get('categoria')
        valor_previsto = float(request.form.get('valor_previsto'))
        valor_realizado = float(request.form.get('valor_realizado', 0))
        
        orcamento = Orcamento(
            ano=ano, mes=mes, categoria=categoria,
            valor_previsto=valor_previsto, valor_realizado=valor_realizado
        )
        db.session.add(orcamento)
        db.session.commit()
        log_auditoria('Orçamento cadastrado', f'{categoria} - {mes}/{ano}')
        flash('Orçamento cadastrado!', 'success')
        return redirect(url_for('orcamento'))
    
    orcamentos = Orcamento.query.order_by(Orcamento.ano.desc(), Orcamento.mes).all()
    total_previsto = sum(o.valor_previsto for o in orcamentos)
    total_realizado = sum(o.valor_realizado for o in orcamentos)
    
    from datetime import date
    return render_template('orcamento.html', 
                         orcamentos=orcamentos,
                         total_previsto=total_previsto,
                         total_realizado=total_realizado,
                         today=date.today())

@app.route('/ajustes')
@login_required
def ajustes():
    precos = PrecoLeite.query.order_by(PrecoLeite.data_vigencia.desc()).all()
    return render_template('ajustes.html', precos=precos)

@app.route('/ajustes/preco', methods=['POST'])
@login_required
def ajustar_preco():
    if current_user.role not in ['admin', 'gerente']:
        flash('Acesso restrito', 'danger')
        return redirect(url_for('ajustes'))
    
    preco_litro = float(request.form.get('preco_litro'))
    data_vigencia = datetime.strptime(request.form.get('data_vigencia'), '%Y-%m-%d').date()
    
    novo_preco = PrecoLeite(preco_litro=preco_litro, data_vigencia=data_vigencia)
    db.session.add(novo_preco)
    db.session.commit()
    log_auditoria('Preço ajustado', f'R$ {preco_litro}/L a partir de {data_vigencia}')
    flash('Preço atualizado!', 'success')
    return redirect(url_for('ajustes'))

# ========== ADMIN ==========

@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    
    usuarios = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin_usuarios.html', usuarios=usuarios)

@app.route('/admin/usuario/criar', methods=['POST'])
@login_required
def admin_usuario_criar():
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    from werkzeug.security import generate_password_hash
    nome = request.form.get('nome')
    email = request.form.get('email')
    senha = request.form.get('senha')
    is_admin = request.form.get('is_admin') == 'on'
    if User.query.filter_by(email=email).first():
        flash('Email já cadastrado', 'danger')
        return redirect(url_for('admin_usuarios'))
    user = User(nome=nome, email=email, senha_hash=generate_password_hash(senha), is_admin=is_admin, role='admin' if is_admin else 'operador')
    db.session.add(user)
    db.session.commit()
    log_auditoria('Usuário criado', f'{email}')
    flash('Usuário criado!', 'success')
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/usuario/editar/<int:id>', methods=['POST'])
@login_required
def admin_usuario_editar(id):
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    from werkzeug.security import generate_password_hash
    user = User.query.get_or_404(id)
    user.nome = request.form.get('nome')
    user.email = request.form.get('email')
    senha = request.form.get('senha')
    if senha:
        user.senha_hash = generate_password_hash(senha)
    user.is_admin = request.form.get('is_admin') == 'on'
    user.role = 'admin' if user.is_admin else 'operador'
    db.session.commit()
    log_auditoria('Usuário editado', f'{user.email}')
    flash('Usuário atualizado!', 'success')
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/usuario/trocar-senha/<int:id>', methods=['POST'])
@login_required
def admin_usuario_trocar_senha(id):
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    from werkzeug.security import generate_password_hash
    user = User.query.get_or_404(id)
    nova_senha = request.form.get('nova_senha')
    user.senha_hash = generate_password_hash(nova_senha)
    db.session.commit()
    log_auditoria('Senha alterada', f'{user.email}')
    flash('Senha alterada!', 'success')
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/usuario/excluir/<int:id>')
@login_required
def admin_usuario_excluir(id):
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('Você não pode excluir a si mesmo!', 'danger')
        return redirect(url_for('admin_usuarios'))
    db.session.delete(user)
    db.session.commit()
    log_auditoria('Usuário excluído', f'{user.email}')
    flash('Usuário excluído!', 'success')
    return redirect(url_for('admin_usuarios'))

@app.route('/admin/auditoria')
@login_required
def auditoria():
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(100).all()
    return render_template('auditoria.html', logs=logs)

# ========== BACKUP ==========

@app.route('/admin/backup')
@login_required
def backup():
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    
    from datetime import datetime
    backup_name = f'backup_terra_roxa_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    
    backup_data = {}
    for table_name in ['animal', 'tipo_racao', 'producao_leite', 'consumo_racao',
                       'preco_leite', 'despesa', 'orcamento', 'user', 'audit_log',
                       'cliente', 'venda_avulsa', 'saude_animal']:
        try:
            table = db.metadata.tables.get(table_name)
            if table is None:
                continue
            rows = db.session.execute(table.select()).fetchall()
            backup_data[table_name] = [dict(row._mapping) for row in rows]
        except Exception:
            backup_data[table_name] = []
    
    from backup_util import BACKUP_DIR
    BACKUP_DIR.mkdir(exist_ok=True)
    filepath = BACKUP_DIR / backup_name
    import json
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2, default=str)
    
    response = make_response(json.dumps(backup_data, ensure_ascii=False, indent=2, default=str))
    response.headers['Content-Disposition'] = f'attachment; filename={backup_name}'
    response.headers['Content-Type'] = 'application/json'
    log_auditoria('Backup realizado', f'Arquivo: {backup_name}')
    return response

@app.route('/admin/backup/listar')
@login_required
def listar_backups():
    if current_user.role != 'admin':
        return jsonify([])
    from backup_util import list_backups
    backups = list_backups()
    return jsonify([{
        'filename': b['filename'],
        'size': f"{b['size'] / 1024:.1f} KB",
        'date': b['date'].strftime('%d/%m/%Y %H:%M')
    } for b in backups])

@app.route('/admin/backup/download/<filename>')
@login_required
def download_backup(filename):
    if current_user.role != 'admin':
        flash('Acesso restrito', 'danger')
        return redirect(url_for('index'))
    from backup_util import BACKUP_DIR
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        flash('Arquivo nao encontrado', 'danger')
        return redirect(url_for('backup'))
    return send_file(str(filepath), as_attachment=True, download_name=filename)

# ========== VENDAS AVULSAS ==========

@app.route('/vendas-avulsas', methods=['GET', 'POST'])
@login_required
def vendas_avulsas():
    from datetime import date
    if request.method == 'POST':
        acao = request.form.get('acao')

        if acao == 'cliente':
            nome = request.form.get('nome')
            if Cliente.query.filter_by(nome=nome).first():
                flash('Cliente já cadastrado!', 'danger')
                return redirect(url_for('vendas_avulsas'))
            cliente = Cliente(nome=nome)
            db.session.add(cliente)
            db.session.commit()
            log_auditoria('Cliente cadastrado', f'{nome}')
            flash('Cliente cadastrado com sucesso!', 'success')
            return redirect(url_for('vendas_avulsas'))

        elif acao == 'venda':
            cliente_id = request.form.get('cliente_id')
            data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
            litros = float(request.form.get('litros'))
            valor_litro = float(request.form.get('valor_litro'))
            total = round(litros * valor_litro, 2)
            venda = VendaAvulsa(cliente_id=cliente_id, data=data, litros=litros, valor_litro=valor_litro, total=total)
            db.session.add(venda)
            db.session.commit()
            log_auditoria('Venda avulsa registrada', f'{litros}L - R$ {total}')
            flash('Venda registrada com sucesso!', 'success')
            return redirect(url_for('vendas_avulsas'))

    clientes = Cliente.query.order_by(Cliente.nome).all()
    vendas = VendaAvulsa.query.order_by(VendaAvulsa.data.desc()).all()
    total_litros_geral = sum(float(v.litros) for v in vendas)
    total_valor_geral = sum(float(v.total) for v in vendas)
    return render_template('vendas_avulsas.html', clientes=clientes, vendas=vendas,
                           total_litros_geral=total_litros_geral,
                           total_valor_geral=total_valor_geral,
                           today=date.today())

@app.route('/vendas-avulsas/exportar/pdf')
@login_required
def vendas_exportar_pdf():
    from datetime import datetime
    vendas = VendaAvulsa.query.order_by(VendaAvulsa.data.asc()).all()
    total_litros = sum(float(v.litros) for v in vendas)
    total_valor = sum(float(v.total) for v in vendas)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=20, textColor=colors.HexColor('#27AE60'),
                                     spaceAfter=20, alignment=TA_CENTER)
        elements.append(Paragraph('Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Vendas Avulsas - {datetime.now().strftime("%d/%m/%Y")}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))

        elements.append(Paragraph(f'<b>Total de Vendas:</b> {len(vendas)}  |  <b>Total Litros:</b> {total_litros:.0f} L  |  <b>Total R$:</b> {total_valor:.2f}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        data = [['Data', 'Cliente', 'Litros', 'R$/L', 'Total']]
        for v in vendas:
            data.append([v.data.strftime('%d/%m/%Y'), v.cliente.nome, f'{v.litros:.1f}', f'{v.valor_litro:.2f}', f'{v.total:.2f}'])

        t = Table(data, colWidths=[3*cm, 5*cm, 2.5*cm, 2.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9F7')])
        ]))
        elements.append(t)

        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=vendas_avulsas_{datetime.now().strftime("%Y%m%d")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        html = '<html><head><title>Vendas Avulsas</title><style>'
        html += 'body{font-family:Arial;padding:40px} h1{color:#27AE60} '
        html += 'table{border-collapse:collapse;width:100%;margin:20px 0} '
        html += 'th,td{border:1px solid #ddd;padding:12px;text-align:center} th{background:#27AE60;color:#fff}'
        html += '</style></head><body>'
        html += f'<h1>Terra Roxa System</h1><h3>Vendas Avulsas</h3>'
        html += f'<p>Total: {len(vendas)} vendas | {total_litros:.0f} L | R$ {total_valor:.2f}</p><hr>'
        html += '<table><tr><th>Data</th><th>Cliente</th><th>Litros</th><th>R$/L</th><th>Total</th></tr>'
        for v in vendas:
            html += f'<tr><td>{v.data.strftime("%d/%m/%Y")}</td><td>{v.cliente.nome}</td><td>{v.litros:.1f}</td><td>{v.valor_litro:.2f}</td><td>R$ {v.total:.2f}</td></tr>'
        html += '</table></body></html>'
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=vendas_avulsas_{datetime.now().strftime("%Y%m%d")}.html'
        response.headers['Content-Type'] = 'text/html'
        return response

@app.route('/vendas-avulsas/exportar/excel')
@login_required
def vendas_exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from io import BytesIO

    vendas = VendaAvulsa.query.order_by(VendaAvulsa.data.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas Avulsas'

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(1, 1, 'Terra Roxa System - Vendas Avulsas').font = Font(bold=True, size=14, color='27AE60')
    ws.merge_cells('A1:E1')
    ws.cell(2, 1, f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    ws.merge_cells('A2:E2')

    headers = ['Data', 'Cliente', 'Litros', 'Valor/Litro', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, v in enumerate(vendas, 5):
        ws.cell(i, 1, v.data.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(i, 2, v.cliente.nome).border = thin_border
        ws.cell(i, 3, float(v.litros)).border = thin_border
        ws.cell(i, 4, float(v.valor_litro)).border = thin_border
        ws.cell(i, 5, float(v.total)).border = thin_border

    row = 5 + len(vendas)
    ws.cell(row, 1, 'TOTAIS').font = Font(bold=True)
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 3, sum(float(v.litros) for v in vendas)).font = Font(bold=True)
    ws.cell(row, 3).border = thin_border
    ws.cell(row, 4).border = thin_border
    ws.cell(row, 5, sum(float(v.total) for v in vendas)).font = Font(bold=True)
    ws.cell(row, 5).border = thin_border

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=vendas_avulsas_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@app.route('/vendas-avulsas/excluir-cliente/<int:id>')
@login_required
def excluir_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    VendaAvulsa.query.filter_by(cliente_id=id).delete()
    db.session.delete(cliente)
    db.session.commit()
    flash('Cliente excluído!', 'success')
    return redirect(url_for('vendas_avulsas'))

@app.route('/vendas-avulsas/editar/<int:id>', methods=['POST'])
@login_required
def editar_venda_avulsa(id):
    venda = VendaAvulsa.query.get_or_404(id)
    venda.cliente_id = request.form.get('cliente_id')
    venda.data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
    venda.valor_litro = float(request.form.get('valor_litro'))
    venda.litros = float(request.form.get('litros'))
    venda.total = round(venda.litros * venda.valor_litro, 2)
    db.session.commit()
    log_auditoria('Venda editada', f'{venda.litros}L - R$ {venda.total}')
    flash('Venda atualizada!', 'success')
    return redirect(url_for('vendas_avulsas'))

@app.route('/vendas-avulsas/excluir/<int:id>')
@login_required
def excluir_venda_avulsa(id):
    venda = VendaAvulsa.query.get(id)
    db.session.delete(venda)
    db.session.commit()
    flash('Venda excluída!', 'success')
    return redirect(url_for('vendas_avulsas'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)
