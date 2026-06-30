from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response, jsonify
from flask_login import login_required, current_user
from extensions import db
from models import Cliente, VendaAvulsa
from utils import log_auditoria
from datetime import datetime, date
from io import BytesIO
from sqlalchemy import func as sa_func
from validation import validate_required, validate_nome, validate_email, validate_telefone, validate_positive, validate_date_range, validate_maxlen

vendas_bp = Blueprint('vendas', __name__)


@vendas_bp.route('/vendas-avulsas', methods=['GET', 'POST'])
@login_required
def vendas_avulsas():
    from datetime import date
    if request.method == 'POST':
        acao = request.form.get('acao')

        if acao == 'cliente':
            nome = request.form.get('nome')
            erro = validate_nome(nome)
            if not erro:
                erro = validate_required(request.form.get('telefone'), 'Telefone')
            if not erro:
                erro = validate_telefone(request.form.get('telefone'))
            if erro:
                flash(erro, 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            email_err = validate_email(request.form.get('email'))
            if email_err:
                flash(email_err, 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            if Cliente.query.filter_by(nome=nome).first():
                flash('Cliente já cadastrado!', 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            endereco = request.form.get('endereco', '')
            erro_addr = validate_maxlen(endereco, 'Endereco', 300)
            if erro_addr:
                flash(erro_addr, 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            cliente = Cliente(
                nome=nome,
                telefone=request.form.get('telefone'),
                email=request.form.get('email'),
                endereco=endereco
            )
            db.session.add(cliente)
            db.session.commit()
            log_auditoria('Cliente cadastrado', f'{nome}')
            flash('Cliente cadastrado com sucesso!', 'success')
            return redirect(url_for('vendas.vendas_avulsas'))

        elif acao == 'venda':
            cliente_id = request.form.get('cliente_id')
            if not cliente_id:
                flash('Selecione um cliente.', 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            data_str = request.form.get('data')
            erro = validate_date_range(data_str)
            if erro:
                flash(erro, 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            data = datetime.strptime(data_str, '%Y-%m-%d').date()
            litros_str = request.form.get('litros')
            erro = validate_required(litros_str, 'Litros')
            if not erro:
                erro = validate_positive(litros_str, 'Litros')
            if erro:
                flash(erro, 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            litros = float(litros_str)
            valor_litro_str = request.form.get('valor_litro')
            erro = validate_required(valor_litro_str, 'Valor por litro')
            if not erro:
                erro = validate_positive(valor_litro_str, 'Valor por litro')
            if erro:
                flash(erro, 'danger')
                return redirect(url_for('vendas.vendas_avulsas'))
            valor_litro = float(valor_litro_str)
            total = round(litros * valor_litro, 2)
            venda = VendaAvulsa(cliente_id=cliente_id, data=data, litros=litros, valor_litro=valor_litro, total=total)
            db.session.add(venda)
            db.session.commit()
            log_auditoria('Venda avulsa registrada', f'{litros}L - R$ {total}')
            flash('Venda registrada com sucesso!', 'success')
            return redirect(url_for('vendas.vendas_avulsas'))

    busca = request.args.get('busca', '').strip()
    clientes = Cliente.query.order_by(Cliente.nome).all()
    vendas_query = VendaAvulsa.query.join(Cliente)
    if busca:
        vendas_query = vendas_query.filter(Cliente.nome.ilike(f'%{busca}%'))
    page = request.args.get('page', 1, type=int)
    vendas_paginator = vendas_query.order_by(VendaAvulsa.data.desc()).paginate(page=page, per_page=50, error_out=False)
    vendas = vendas_paginator.items
    total_litros_geral = vendas_query.with_entities(sa_func.coalesce(sa_func.sum(VendaAvulsa.litros), 0)).scalar()
    total_valor_geral = vendas_query.with_entities(sa_func.coalesce(sa_func.sum(VendaAvulsa.total), 0)).scalar()
    return render_template('vendas_avulsas.html', clientes=clientes, vendas=vendas,
                           paginator=vendas_paginator,
                           total_litros_geral=total_litros_geral,
                           total_valor_geral=total_valor_geral,
                           today=date.today(), busca=busca)


@vendas_bp.route('/vendas-avulsas/exportar/pdf')
@login_required
def vendas_exportar_pdf():
    from datetime import datetime
    vendas = VendaAvulsa.query.order_by(VendaAvulsa.data.asc()).all()
    total_litros = sum(float(v.litros) for v in vendas)
    total_valor = sum(float(v.total) for v in vendas)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=20, textColor=colors.HexColor('#27AE60'),
                                     spaceAfter=20, alignment=TA_CENTER)
        elements.append(Paragraph('Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Vendas Avulsas - {datetime.now().strftime("%d/%m/%Y")}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))

        elements.append(Paragraph(f'<b>Total de Vendas:</b> {len(vendas)}  |  <b>Total Litros:</b> {total_litros:.0f} L  |  <b>Total R$:</b> {total_valor:.2f}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        data = [['Data', 'Cliente', 'Litros', 'R$/L', 'Total']]
        for v in vendas:
            data.append([v.data.strftime('%d/%m/%Y'), v.cliente.nome, f'{v.litros:.1f}', f'{v.valor_litro:.2f}', f'{v.total:.2f}'])

        t = Table(data, colWidths=[3*cm, 5*cm, 2.5*cm, 2.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9F7')])
        ]))
        elements.append(t)

        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=vendas_avulsas_{datetime.now().strftime("%Y%m%d")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        html = '<html><head><title>Vendas Avulsas</title><style>'
        html += 'body{font-family:Arial;padding:40px} h1{color:#27AE60} '
        html += 'table{border-collapse:collapse;width:100%;margin:20px 0} '
        html += 'th,td{border:1px solid #ddd;padding:12px;text-align:center} th{background:#27AE60;color:#fff}'
        html += '</style></head><body>'
        html += f'<h1>Terra Roxa System</h1><h3>Vendas Avulsas</h3>'
        html += f'<p>Total: {len(vendas)} vendas | {total_litros:.0f} L | R$ {total_valor:.2f}</p><hr>'
        html += '<table><tr><th>Data</th><th>Cliente</th><th>Litros</th><th>R$/L</th><th>Total</th></tr>'
        for v in vendas:
            html += f'<tr><td>{v.data.strftime("%d/%m/%Y")}</td><td>{v.cliente.nome}</td><td>{v.litros:.1f}</td><td>{v.valor_litro:.2f}</td><td>R$ {v.total:.2f}</td></tr>'
        html += '</table></body></html>'
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=vendas_avulsas_{datetime.now().strftime("%Y%m%d")}.html'
        response.headers['Content-Type'] = 'text/html'
        return response


@vendas_bp.route('/vendas-avulsas/exportar/excel')
@login_required
def vendas_exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from io import BytesIO

    vendas = VendaAvulsa.query.order_by(VendaAvulsa.data.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas Avulsas'

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(1, 1, 'Terra Roxa System - Vendas Avulsas').font = Font(bold=True, size=14, color='27AE60')
    ws.merge_cells('A1:E1')
    ws.cell(2, 1, f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    ws.merge_cells('A2:E2')

    headers = ['Data', 'Cliente', 'Litros', 'Valor/Litro', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, v in enumerate(vendas, 5):
        ws.cell(i, 1, v.data.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(i, 2, v.cliente.nome).border = thin_border
        ws.cell(i, 3, float(v.litros)).border = thin_border
        ws.cell(i, 4, float(v.valor_litro)).border = thin_border
        ws.cell(i, 5, float(v.total)).border = thin_border

    row = 5 + len(vendas)
    ws.cell(row, 1, 'TOTAIS').font = Font(bold=True)
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 3, sum(float(v.litros) for v in vendas)).font = Font(bold=True)
    ws.cell(row, 3).border = thin_border
    ws.cell(row, 4).border = thin_border
    ws.cell(row, 5, sum(float(v.total) for v in vendas)).font = Font(bold=True)
    ws.cell(row, 5).border = thin_border

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=vendas_avulsas_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@vendas_bp.route('/vendas-avulsas/excluir-cliente/<int:id>')
@login_required
def excluir_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    try:
        VendaAvulsa.query.filter_by(cliente_id=id).delete()
        db.session.delete(cliente)
        db.session.commit()
        flash('Cliente excluído!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir cliente: {str(e)}', 'danger')
    return redirect(url_for('vendas.vendas_avulsas'))


@vendas_bp.route('/vendas-avulsas/editar/<int:id>', methods=['POST'])
@login_required
def editar_venda_avulsa(id):
    venda = VendaAvulsa.query.get_or_404(id)
    try:
        venda.cliente_id = request.form.get('cliente_id')
        venda.data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        venda.valor_litro = float(request.form.get('valor_litro'))
        venda.litros = float(request.form.get('litros'))
        venda.total = round(venda.litros * venda.valor_litro, 2)
        db.session.commit()
        log_auditoria('Venda editada', f'{venda.litros}L - R$ {venda.total}')
        flash('Venda atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar venda: {str(e)}', 'danger')
    return redirect(url_for('vendas.vendas_avulsas'))


@vendas_bp.route('/vendas-avulsas/excluir/<int:id>')
@login_required
def excluir_venda_avulsa(id):
    venda = VendaAvulsa.query.get_or_404(id)
    try:
        db.session.delete(venda)
        db.session.commit()
        flash('Venda excluída!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir venda: {str(e)}', 'danger')
    return redirect(url_for('vendas.vendas_avulsas'))
