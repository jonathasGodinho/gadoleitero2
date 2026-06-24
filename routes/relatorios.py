from flask import Blueprint, render_template, request, make_response, jsonify
from flask_login import login_required
from extensions import db
from models import ProducaoLeite, ConsumoRacao, Despesa, VendaAvulsa
from utils import get_preco_vigente, calcular_custo_producao, evolucao_custo_litro_periodo
from datetime import datetime, date, timedelta
from io import BytesIO

relatorios_bp = Blueprint('relatorios', __name__)

@relatorios_bp.route('/relatorios')
@login_required
def relatorios():
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    tipo = request.args.get('tipo', 'producao')
    
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    if tipo == 'vendas-avulsas':
        vendas = VendaAvulsa.query.filter(
            VendaAvulsa.data >= data_ini_date, VendaAvulsa.data <= data_fim_date
        ).order_by(VendaAvulsa.data.desc()).all()
        
        total_litros_vendas = sum(float(v.litros) for v in vendas)
        total_valor_vendas = sum(float(v.total) for v in vendas)
        
        dias_grafico = []
        valores_venda = []
        delta = data_fim_date - data_ini_date
        if delta.days <= 31:
            for i in range(delta.days + 1):
                d = data_ini_date + timedelta(days=i)
                dias_grafico.append(d.strftime('%d/%m'))
                vals_dia = [float(v.total) for v in vendas if v.data == d]
                valores_venda.append(sum(vals_dia))
        
        vendas_por_mes = {}
        for v in vendas:
            mes_key = v.data.strftime('%Y-%m')
            vendas_por_mes[mes_key] = vendas_por_mes.get(mes_key, 0) + float(v.total)
        meses_vendas = sorted(vendas_por_mes.keys())
        if meses_vendas:
            meses_labels_vendas = [datetime.strptime(m + '-01', '%Y-%m-%d').strftime('%m/%Y') for m in meses_vendas]
            valores_vendas_mensais = [vendas_por_mes[m] for m in meses_vendas]
        else:
            meses_labels_vendas = []
            valores_vendas_mensais = []
        
        return render_template('relatorios.html',
                               data_ini=data_ini, data_fim=data_fim,
                               tipo=tipo,
                               vendas=vendas,
                               total_litros_vendas=total_litros_vendas,
                               total_valor_vendas=total_valor_vendas,
                               dias_grafico=dias_grafico,
                               valores_venda=valores_venda,
                               meses_labels_vendas=meses_labels_vendas,
                               valores_vendas_mensais=valores_vendas_mensais)
    
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini_date, ConsumoRacao.data <= data_fim_date
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini_date, Despesa.data <= data_fim_date
    ).all()
    
    total_litros = sum(p.litros for p in producoes)
    custo_total = sum(c.custo for c in consumos) + sum(d.valor for d in despesas)
    
    receita = 0
    for p in producoes:
        preco = float(p.preco_venda) if p.preco_venda else float(get_preco_vigente(p.data))
        receita += float(p.litros) * preco
    
    lucro = receita - float(custo_total)
    custo_producao = calcular_custo_producao(data_ini_date, data_fim_date)
    
    dias_grafico = []
    valores_prod = []
    custos_dia = []
    
    delta = data_fim_date - data_ini_date
    if delta.days <= 31:
        for i in range(delta.days + 1):
            d = data_ini_date + timedelta(days=i)
            dias_grafico.append(d.strftime('%d/%m'))
            
            prods_dia = [p.litros for p in producoes if p.data == d]
            valores_prod.append(sum(prods_dia))
            
            custos_dia_sum = [c.custo for c in consumos if c.data == d]
            custos_dia.append(sum(custos_dia_sum))
    
    custo_por_tipo = {}
    for c in consumos:
        tipo_nome = c.tipo_racao.nome
        custo_por_tipo[tipo_nome] = custo_por_tipo.get(tipo_nome, 0) + float(c.custo)

    producao_por_mes = {}
    for p in producoes:
        mes_key = p.data.strftime('%Y-%m')
        producao_por_mes[mes_key] = producao_por_mes.get(mes_key, 0) + float(p.litros)

    custo_por_mes = {}
    for c in consumos:
        mes_key = c.data.strftime('%Y-%m')
        custo_por_mes[mes_key] = custo_por_mes.get(mes_key, 0) + float(c.custo)
    for d in despesas:
        mes_key = d.data.strftime('%Y-%m')
        custo_por_mes[mes_key] = custo_por_mes.get(mes_key, 0) + float(d.valor)

    receita_por_mes = {}
    for p in producoes:
        mes_key = p.data.strftime('%Y-%m')
        preco = float(p.preco_venda) if p.preco_venda else float(get_preco_vigente(p.data))
        receita_por_mes[mes_key] = receita_por_mes.get(mes_key, 0) + float(p.litros) * preco

    todos_meses = sorted(set(list(producao_por_mes.keys()) + list(custo_por_mes.keys()) + list(receita_por_mes.keys())))
    if todos_meses:
        meses_labels = [datetime.strptime(m + '-01', '%Y-%m-%d').strftime('%m/%Y') for m in todos_meses]
        valores_mensais = [producao_por_mes.get(m, 0) for m in todos_meses]
        custo_mensal = [custo_por_mes.get(m, 0) for m in todos_meses]
        lucro_mensal = [receita_por_mes.get(m, 0) - custo_por_mes.get(m, 0) for m in todos_meses]
    else:
        meses_labels = []
        valores_mensais = []
        custo_mensal = []
        lucro_mensal = []
    
    preco_medio = get_preco_vigente(data_fim_date)

    rel_custo_litro_dias, rel_custo_litro_valores = evolucao_custo_litro_periodo(data_ini_date, data_fim_date)
    rel_custo_litro_media = round(sum(rel_custo_litro_valores) / len(rel_custo_litro_valores), 4) if rel_custo_litro_valores else 0

    return render_template('relatorios.html',
                           data_ini=data_ini, data_fim=data_fim,
                           tipo=tipo,
                           total_litros=total_litros, receita=receita,
                           custo=custo_total, lucro=lucro,
                           custo_producao=custo_producao,
                           dias_grafico=dias_grafico, valores_prod=valores_prod,
                           custos_dia=custos_dia, custo_por_tipo=custo_por_tipo,
                           preco_medio=preco_medio,
                           meses_labels=meses_labels, valores_mensais=valores_mensais,
                           custo_mensal=custo_mensal, lucro_mensal=lucro_mensal,
                           rel_custo_litro_dias=rel_custo_litro_dias,
                           rel_custo_litro_valores=rel_custo_litro_valores,
                           rel_custo_litro_media=rel_custo_litro_media)

@relatorios_bp.route('/relatorios/pdf')
@login_required
def relatorio_pdf():
    from datetime import datetime
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini_date, ConsumoRacao.data <= data_fim_date
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini_date, Despesa.data <= data_fim_date
    ).all()
    
    total_litros = sum(p.litros for p in producoes)
    custo_total = sum(c.custo for c in consumos) + sum(d.valor for d in despesas)
    receita = sum(float(p.litros) * float(p.preco_venda if p.preco_venda else get_preco_vigente(p.data)) for p in producoes)
    lucro = receita - float(custo_total)
    
    # Tentar usar reportlab para PDF profissional
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.graphics.charts.lineplots import LinePlot
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics import renderPDF
        import tempfile
        import os
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#27AE60'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Cabeçalho
        elements.append(Paragraph('🐄 Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Gestão - {data_ini} a {data_fim}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Linha decorativa
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))
        
        # Resumo Executivo
        elements.append(Paragraph('Resumo Executivo', styles['Heading2']))
        
        # Tabela de métricas
        data = [
            ['Métrica', 'Valor', 'Detalhes'],
            ['Total Litros', f'{total_litros:.2f} L', f'{total_litros/max(1, (data_fim_date-data_ini_date).days):.2f} L/dia'],
            ['Receita Total', f'R$ {receita:.2f}', f'Média: R$ {receita/max(1,total_litros):.4f}/L'],
            ['Custo Total', f'R$ {custo_total:.2f}', f'{(custo_total/receita*100) if receita > 0 else 0:.1f}% da receita'],
            ['Lucro Líquido', f'R$ {lucro:.2f}', f'Margem: {(lucro/receita*100) if receita > 0 else 0:.1f}%']
        ]
        
        t = Table(data, colWidths=[4*cm, 4*cm, 6*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9F7')])
        ]))
        elements.append(t)
        elements.append(Spacer(1, 1*cm))
        
        # Produção por Animal (Top 5)
        if producoes:
            from collections import defaultdict
            prod_animal = defaultdict(float)
            for p in producoes:
                nome_animal = p.animal.nome if p.animal else 'Produção Geral'
                prod_animal[nome_animal] += float(p.litros)
            
            top5 = sorted(prod_animal.items(), key=lambda x: x[1], reverse=True)[:5]
            
            elements.append(Paragraph('Top 5 Animais - Produção', styles['Heading2']))
            
            data_animais = [['Animal', 'Total (L)', '% do Total']]
            for animal, prod in top5:
                pct = (prod / total_litros * 100) if total_litros > 0 else 0
                data_animais.append([animal, f'{prod:.2f}', f'{pct:.1f}%'])
            
            t2 = Table(data_animais, colWidths=[6*cm, 4*cm, 4*cm])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF5FB')])
            ]))
            elements.append(t2)
            elements.append(Spacer(1, 1*cm))
        
        # Custos por Categoria
        if despesas:
            from collections import defaultdict
            custos_cat = defaultdict(float)
            for d in despesas:
                custos_cat[d.categoria or 'Outros'] += float(d.valor)
            
            elements.append(Paragraph('Custos por Categoria', styles['Heading2']))
            
            data_custos = [['Categoria', 'Valor (R$)', '% do Total']]
            for cat, val in sorted(custos_cat.items(), key=lambda x: x[1], reverse=True):
                pct = (val / custo_total * 100) if custo_total > 0 else 0
                data_custos.append([cat, f'{val:.2f}', f'{pct:.1f}%'])
            
            t3 = Table(data_custos, colWidths=[6*cm, 4*cm, 4*cm])
            t3.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F39C12')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF9E7')])
            ]))
            elements.append(t3)
        
        # Rodapé
        elements.append(Spacer(1, 2*cm))
        elements.append(Paragraph(f'Relatório gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', 
                                            styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_terra_roxa_{data_ini}_{data_fim}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
        
    except Exception as e:
        # Fallback para HTML
        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>Relatório Terra Roxa</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 40px; }}
                h1 {{ color: #27AE60; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: center; }}
                th {{ background: #27AE60; color: white; }}
                .positive {{ color: green; }}
                .negative {{ color: red; }}
            </style>
        </head>
        <body>
            <h1>🐄 Terra Roxa System</h1>
            <h3>Relatório de Gestão</h3>
            <p><strong>Período:</strong> {data_ini} a {data_fim}</p>
            <hr>
            <h4>Resumo Executivo</h4>
            <table>
                <tr><th>Métrica</th><th>Valor</th></tr>
                <tr><td>Total Litros</td><td>{total_litros:.2f} L</td></tr>
                <tr><td>Receita Total</td><td>R$ {receita:.2f}</td></tr>
                <tr><td>Custo Total</td><td>R$ {custo_total:.2f}</td></tr>
                <tr><td>Lucro Líquido</td><td class="{'positive' if lucro >= 0 else 'negative'}">R$ {lucro:.2f}</td></tr>
            </table>
            <p style="margin-top: 40px; font-size: 12px; color: #666;">
                Relatório gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System
            </p>
        </body>
        </html>
        '''
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=relatorio_terra_roxa_{data_ini}_{data_fim}.html'
        response.headers['Content-Type'] = 'text/html'
        return response

@relatorios_bp.route('/relatorios/excel')
@login_required
def relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from io import BytesIO
    from decimal import Decimal

    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')

    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')

    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()

    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).all()
    consumos = ConsumoRacao.query.filter(
        ConsumoRacao.data >= data_ini_date, ConsumoRacao.data <= data_fim_date
    ).all()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini_date, Despesa.data <= data_fim_date
    ).all()

    total_litros = sum(p.litros for p in producoes)
    custo_total = float(sum(c.custo for c in consumos)) + float(sum(d.valor for d in despesas))
    receita = sum(float(p.litros) * float(p.preco_venda if p.preco_venda else get_preco_vigente(p.data)) for p in producoes)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Resumo ---
    ws = wb.active
    ws.title = 'Resumo'
    ws.cell(1, 1, 'Terra Roxa System - Relatorio de Gestao').font = Font(bold=True, size=14, color='27AE60')
    ws.merge_cells('A1:B1')
    ws.cell(2, 1, f'Periodo: {data_ini} a {data_fim}')
    ws.merge_cells('A2:B2')

    headers = ['Metrica', 'Valor']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    rows = [
        ('Total Litros', f'{total_litros:.2f} L'),
        ('Receita Total', f'R$ {receita:.2f}'),
        ('Custo Total', f'R$ {custo_total:.2f}'),
        ('Lucro Liquido', f'R$ {receita - custo_total:.2f}'),
    ]
    for i, (k, v) in enumerate(rows, 5):
        ws.cell(i, 1, k).border = thin_border
        ws.cell(i, 2, v).border = thin_border

    # --- Producao ---
    ws2 = wb.create_sheet('Producao')
    headers2 = ['Data', 'Animal', 'Litros', 'Preco/L', 'Valor']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, p in enumerate(producoes, 2):
        prec = p.preco_venda if p.preco_venda else get_preco_vigente(p.data)
        nome = p.animal.nome if p.animal else 'Geral'
        ws2.cell(i, 1, p.data.isoformat()).border = thin_border
        ws2.cell(i, 2, nome).border = thin_border
        ws2.cell(i, 3, float(p.litros)).border = thin_border
        ws2.cell(i, 4, float(prec)).border = thin_border
        ws2.cell(i, 5, round(float(p.litros) * float(prec), 2)).border = thin_border

    # --- Custos ---
    ws3 = wb.create_sheet('Custos')
    headers3 = ['Tipo', 'Data', 'Categoria/Descricao', 'Valor']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    i = 2
    for c in consumos:
        ws3.cell(i, 1, 'Racao').border = thin_border
        ws3.cell(i, 2, c.data.isoformat()).border = thin_border
        ws3.cell(i, 3, c.tipo_racao.nome).border = thin_border
        ws3.cell(i, 4, float(c.custo)).border = thin_border
        i += 1
    for d in despesas:
        ws3.cell(i, 1, 'Despesa').border = thin_border
        ws3.cell(i, 2, d.data.isoformat()).border = thin_border
        ws3.cell(i, 3, d.categoria or 'Geral').border = thin_border
        ws3.cell(i, 4, float(d.valor)).border = thin_border
        i += 1

    for ws_ in [ws, ws2, ws3]:
        for col in range(1, 6):
            letter = chr(64 + col)
            ws_.column_dimensions[letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_terra_roxa_{data_ini}_{data_fim}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

@relatorios_bp.route('/cooperativas')
@login_required
def cooperativas():
    # Integração com cooperativas (simulado por enquanto)
    # Em produção, usaria APIs reais de cooperativas
    cooperativas_info = [
        {
            'nome': 'Coopela',
            'contato': '(11) 1234-5678',
            'email': 'contato@coopela.com.br',
            'preco_leite': 3.60,
            'cargas_disponiveis': 5,
            'distancia_km': 45
        },
        {
            'nome': 'Laticínios Boa Vista',
            'contato': '(11) 9876-5432',
            'email': 'comercial@laticinios.com.br',
            'preco_leite': 3.55,
            'cargas_disponiveis': 3,
            'distancia_km': 60
        }
    ]
    return render_template('cooperativas.html', cooperativas=cooperativas_info)
