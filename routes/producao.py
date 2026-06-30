from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from extensions import db
from models import ProducaoLeite, Animal
from utils import log_auditoria, get_preco_vigente
from datetime import datetime, date, timedelta
from io import BytesIO
import calendar
from sqlalchemy import func as sa_func
from validation import validate_positive, validate_date_range, validate_maxlen, validate_required

producao_bp = Blueprint('producao', __name__)


@producao_bp.route('/producao', methods=['GET', 'POST'])
@login_required
def producao():
    if request.method == 'POST':
        erro = validate_required(request.form.get('litros'), 'Litros')
        if not erro:
            erro = validate_positive(request.form.get('litros'), 'Litros')
        if erro:
            flash(erro, 'danger')
            return redirect(url_for('producao.producao'))
        litros = float(request.form.get('litros'))

        animal_id = request.form.get('animal_id')
        erro = validate_required(animal_id, 'Animal')
        if erro:
            flash(erro, 'danger')
            return redirect(url_for('producao.producao'))
        animal = Animal.query.get(int(animal_id))
        if not animal or not animal.ativo:
            flash('Animal inválido ou inativo', 'danger')
            return redirect(url_for('producao.producao'))

        data_str = request.form.get('data')
        erro = validate_required(data_str, 'Data')
        if not erro:
            erro = validate_date_range(data_str)
        if erro:
            flash(erro, 'danger')
            return redirect(url_for('producao.producao'))
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
        try:
            preco_venda = float(request.form.get('preco_venda') or 0)
        except (ValueError, TypeError):
            preco_venda = float(get_preco_vigente(data))
        total_receber = round(litros * preco_venda, 2)
        nova_producao = ProducaoLeite(
            animal_id=animal.id, litros=litros, data=data, 
            preco_venda=preco_venda,
            total_receber=total_receber
        )
        try:
            db.session.add(nova_producao)
            db.session.commit()
            log_auditoria('Produção registrada', f'{animal.nome} - {litros}L a R$ {preco_venda}/L')
            flash('Produção registrada com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao registrar produção: {str(e)}', 'danger')
        return redirect(url_for('producao.producao'))
    
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    ultimo_dia_mes = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1])
    
    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    
    if filtro_data_ini:
        data_ini = datetime.strptime(filtro_data_ini, '%Y-%m-%d').date()
    else:
        data_ini = primeiro_dia_mes
        filtro_data_ini = data_ini.strftime('%Y-%m-%d')
    
    if filtro_data_fim:
        data_fim = datetime.strptime(filtro_data_fim, '%Y-%m-%d').date()
    else:
        data_fim = ultimo_dia_mes
        filtro_data_fim = data_fim.strftime('%Y-%m-%d')
    
    busca = request.args.get('busca', '').strip()
    producoes_query = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini,
        ProducaoLeite.data <= data_fim
    )
    if busca:
        producoes_query = producoes_query.join(Animal).filter(Animal.nome.ilike(f'%{busca}%'))
    page = request.args.get('page', 1, type=int)
    producoes_paginator = producoes_query.order_by(ProducaoLeite.data.desc()).paginate(page=page, per_page=50, error_out=False)
    producoes = producoes_paginator.items
    
    total_litros_geral = producoes_query.with_entities(sa_func.coalesce(sa_func.sum(ProducaoLeite.litros), 0)).scalar()
    total_receber_geral = producoes_query.with_entities(sa_func.coalesce(sa_func.sum(ProducaoLeite.total_receber), 0)).scalar()
    
    today = hoje.strftime('%Y-%m-%d')
    animais = Animal.query.filter_by(ativo=True).order_by(Animal.nome).all()
    
    return render_template('producao.html', 
                           producoes=producoes,
                           paginator=producoes_paginator,
                           today=today,
                           animais=animais,
                           preco_padrao=float(get_preco_vigente(hoje)),
                           total_litros_geral=total_litros_geral,
                           total_receber_geral=total_receber_geral,
                           data_ini=filtro_data_ini,
                           data_fim=filtro_data_fim,
                           busca=busca)


@producao_bp.route('/producao/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producao(id):
    producao = ProducaoLeite.query.get_or_404(id)
    if request.method == 'POST':
        try:
            producao.animal_id = int(request.form.get('animal_id'))
            producao.data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
            producao.litros = float(request.form.get('litros'))
            producao.preco_venda = float(request.form.get('preco_venda'))
            producao.total_receber = round(producao.litros * producao.preco_venda, 2)
            db.session.commit()
            log_auditoria('Produção editada', f'{producao.litros}L a R$ {producao.preco_venda}/L')
            flash('Produção atualizada!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao editar produção: {str(e)}', 'danger')
        return redirect(url_for('producao.producao'))
    animais = Animal.query.filter_by(ativo=True).order_by(Animal.nome).all()
    return render_template('editar_producao.html', producao=producao, animais=animais)


@producao_bp.route('/producao/excluir/<int:id>')
@login_required
def excluir_producao(id):
    if current_user.role not in ['admin', 'gerente']:
        flash('Acesso restrito', 'danger')
        return redirect(url_for('producao.producao'))
    producao = ProducaoLeite.query.get_or_404(id)
    try:
        db.session.delete(producao)
        db.session.commit()
        log_auditoria('Produção excluída', f'ID {id}')
        flash('Produção excluída!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir produção: {str(e)}', 'danger')
    return redirect(url_for('producao.producao'))


@producao_bp.route('/producao/exportar/pdf')
@login_required
def producao_exportar_pdf():
    from datetime import datetime
    from decimal import Decimal
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).order_by(ProducaoLeite.data.asc()).all()
    total_litros = sum(p.litros for p in producoes)
    total_receber = sum(p.total_receber or 0 for p in producoes)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=20, textColor=colors.HexColor('#27AE60'),
                                     spaceAfter=20, alignment=TA_CENTER)
        elements.append(Paragraph('Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Produção - {data_ini} a {data_fim}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(f'<b>Total de Registros:</b> {len(producoes)}  |  <b>Total Litros:</b> {float(total_litros):.0f} L  |  <b>Total a Receber:</b> R$ {float(total_receber):.2f}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        data_table = [['Data', 'Litros', 'Preço/L', 'Total']]
        for p in producoes:
            data_table.append([
                p.data.strftime('%d/%m/%Y'),
                f'{float(p.litros):.1f}',
                f'R$ {float(p.preco_venda or 0):.2f}',
                f'R$ {float(p.total_receber or 0):.2f}'
            ])
        t = Table(data_table, colWidths=[4*cm, 3*cm, 3*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9F7')])
        ]))
        elements.append(t)
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', styles['Normal']))
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=producao_{data_ini}_{data_fim}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception as e:
        html = f'''
        <html><head><title>Produção</title><style>
        body{{font-family:Arial;padding:40px}} h1{{color:#27AE60}}
        table{{border-collapse:collapse;width:100%;margin:20px 0}}
        th,td{{border:1px solid #ddd;padding:12px;text-align:center}} th{{background:#27AE60;color:#fff}}
        </style></head><body>
        <h1>Terra Roxa System</h1><h3>Relatório de Produção</h3>
        <p><strong>Período:</strong> {data_ini} a {data_fim}</p>
        <p>Total: {len(producoes)} registros | {float(total_litros):.0f} L | R$ {float(total_receber):.2f}</p><hr>
        <table><tr><th>Data</th><th>Litros</th><th>Preço/L</th><th>Total</th></tr>'''
        for p in producoes:
            html += f'<tr><td>{p.data.strftime("%d/%m/%Y")}</td><td>{float(p.litros):.1f}</td><td>R$ {float(p.preco_venda or 0):.2f}</td><td>R$ {float(p.total_receber or 0):.2f}</td></tr>'
        html += '</table></body></html>'
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=producao_{data_ini}_{data_fim}.html'
        response.headers['Content-Type'] = 'text/html'
        return response


@producao_bp.route('/producao/exportar/excel')
@login_required
def producao_exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    from decimal import Decimal
    data_ini = request.args.get('data_ini')
    data_fim = request.args.get('data_fim')
    if not data_ini:
        data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = date.today().strftime('%Y-%m-%d')
    data_ini_date = datetime.strptime(data_ini, '%Y-%m-%d').date()
    data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
    producoes = ProducaoLeite.query.filter(
        ProducaoLeite.data >= data_ini_date, ProducaoLeite.data <= data_fim_date
    ).order_by(ProducaoLeite.data.asc()).all()
    total_litros = sum(p.litros for p in producoes)
    total_receber = sum(p.total_receber or 0 for p in producoes)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Producao'
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(1, 1, 'Terra Roxa System - Relatorio de Producao').font = Font(bold=True, size=14, color='27AE60')
    ws.merge_cells('A1:D1')
    ws.cell(2, 1, f'Periodo: {data_ini} a {data_fim}')
    ws.merge_cells('A2:D2')
    headers = ['Data', 'Litros', 'Preco/L', 'Total a Receber']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for i, p in enumerate(producoes, 5):
        ws.cell(i, 1, p.data.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(i, 2, float(p.litros)).border = thin_border
        ws.cell(i, 3, float(p.preco_venda or 0)).border = thin_border
        ws.cell(i, 4, float(p.total_receber or 0)).border = thin_border
    row = 5 + len(producoes)
    ws.cell(row, 1, 'TOTAIS').font = Font(bold=True)
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 2, float(total_litros)).font = Font(bold=True)
    ws.cell(row, 2).border = thin_border
    ws.cell(row, 3).border = thin_border
    ws.cell(row, 4, float(total_receber)).font = Font(bold=True)
    ws.cell(row, 4).border = thin_border
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 18
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=producao_{data_ini}_{data_fim}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
