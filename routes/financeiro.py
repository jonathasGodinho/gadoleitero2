from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from extensions import db
from models import Despesa, Orcamento
from utils import log_auditoria
from datetime import datetime, date
from io import BytesIO
from sqlalchemy import func as sa_func
from validation import validate_required, validate_positive, validate_date_range, validate_maxlen, validate_descricao

financeiro_bp = Blueprint('financeiro', __name__)


@financeiro_bp.route('/financeiro/editar/<int:id>', methods=['POST'])
@login_required
def editar_despesa(id):
    despesa = Despesa.query.get_or_404(id)
    try:
        despesa.descricao = request.form.get('descricao', despesa.descricao)
        despesa.categoria = request.form.get('categoria', despesa.categoria)
        despesa.valor = float(request.form.get('valor', despesa.valor))
        despesa.data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        despesa.observacoes = request.form.get('observacoes', despesa.observacoes)
        db.session.commit()
        log_auditoria('Despesa editada', f'{despesa.descricao} - R$ {despesa.valor}')
        flash('Despesa atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao editar despesa: {str(e)}', 'danger')
    return redirect(url_for('financeiro.financeiro'))


@financeiro_bp.route('/financeiro/excluir/<int:id>')
@login_required
def excluir_despesa(id):
    despesa = Despesa.query.get_or_404(id)
    try:
        db.session.delete(despesa)
        db.session.commit()
        log_auditoria('Despesa excluída', f'{despesa.descricao} - R$ {despesa.valor}')
        flash('Despesa excluída!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir despesa: {str(e)}', 'danger')
    return redirect(url_for('financeiro.financeiro'))


@financeiro_bp.route('/financeiro', methods=['GET', 'POST'])
@login_required
def financeiro():
    if request.method == 'POST':
        descricao = request.form.get('descricao')
        erro = validate_required(descricao, 'Descricao')
        if not erro:
            erro = validate_descricao(descricao)
        if not erro:
            erro = validate_positive(request.form.get('valor'), 'Valor')
        if not erro:
            erro = validate_date_range(request.form.get('data'))
        if erro:
            flash(erro, 'danger')
            return redirect(url_for('financeiro.financeiro'))
        valor = float(request.form.get('valor'))
        categoria = request.form.get('categoria')
        data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        observacoes = request.form.get('observacoes')

        try:
            despesa = Despesa(
                descricao=descricao, valor=valor, categoria=categoria,
                data=data, observacoes=observacoes
            )
            db.session.add(despesa)
            db.session.commit()
            log_auditoria('Despesa registrada', f'{descricao} - R$ {valor}')
            flash('Despesa registrada!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao registrar despesa: {str(e)}', 'danger')
        return redirect(url_for('financeiro.financeiro'))

    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    filtro_categoria = request.args.get('categoria')
    busca = request.args.get('busca', '').strip()

    query = Despesa.query
    if filtro_data_ini:
        query = query.filter(Despesa.data >= datetime.strptime(filtro_data_ini, '%Y-%m-%d').date())
    if filtro_data_fim:
        query = query.filter(Despesa.data <= datetime.strptime(filtro_data_fim, '%Y-%m-%d').date())
    if filtro_categoria:
        query = query.filter(Despesa.categoria == filtro_categoria)
    if busca:
        like = f'%{busca}%'
        query = query.filter(
            db.or_(
                Despesa.descricao.ilike(like),
                Despesa.categoria.ilike(like),
                Despesa.observacoes.ilike(like)
            )
        )

    page = request.args.get('page', 1, type=int)
    despesas_paginator = query.order_by(Despesa.data.desc()).paginate(page=page, per_page=50, error_out=False)
    despesas = despesas_paginator.items

    total_despesas = query.with_entities(sa_func.coalesce(sa_func.sum(Despesa.valor), 0)).scalar()

    return render_template('financeiro.html', despesas=despesas, total_despesas=total_despesas, today=date.today(), busca=busca, paginator=despesas_paginator)


@financeiro_bp.route('/orcamento', methods=['GET', 'POST'])
@login_required
def orcamento():
    if request.method == 'POST':
        try:
            ano = int(request.form.get('ano'))
            mes = int(request.form.get('mes'))
        except (ValueError, TypeError):
            flash('Ano ou mês inválido', 'danger')
            return redirect(url_for('financeiro.orcamento'))
        if ano < 2000 or ano > 2100 or mes < 1 or mes > 12:
            flash('Ano ou mês fora do intervalo válido', 'danger')
            return redirect(url_for('financeiro.orcamento'))
        erro = validate_positive(request.form.get('valor_previsto'), 'Valor previsto')
        if erro:
            flash(erro, 'danger')
            return redirect(url_for('financeiro.orcamento'))
        categoria = request.form.get('categoria')
        valor_previsto = float(request.form.get('valor_previsto'))
        valor_realizado = float(request.form.get('valor_realizado', 0))

        try:
            orcamento = Orcamento(
                ano=ano, mes=mes, categoria=categoria,
                valor_previsto=valor_previsto, valor_realizado=valor_realizado
            )
            db.session.add(orcamento)
            db.session.commit()
            log_auditoria('Orçamento cadastrado', f'{categoria} - {mes}/{ano}')
            flash('Orçamento cadastrado!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar orçamento: {str(e)}', 'danger')
        return redirect(url_for('financeiro.orcamento'))

    page = request.args.get('page', 1, type=int)
    orc_query = Orcamento.query.order_by(Orcamento.ano.desc(), Orcamento.mes)
    orc_paginator = orc_query.paginate(page=page, per_page=50, error_out=False)
    orcamentos = orc_paginator.items
    total_previsto = orc_query.with_entities(sa_func.coalesce(sa_func.sum(Orcamento.valor_previsto), 0)).scalar()
    total_realizado = orc_query.with_entities(sa_func.coalesce(sa_func.sum(Orcamento.valor_realizado), 0)).scalar()

    return render_template('orcamento.html',
                          orcamentos=orcamentos,
                          paginator=orc_paginator,
                          total_previsto=total_previsto,
                          total_realizado=total_realizado,
                          today=date.today())


@financeiro_bp.route('/orcamento/excluir/<int:id>')
@login_required
def excluir_orcamento(id):
    orcamento = Orcamento.query.get_or_404(id)
    try:
        db.session.delete(orcamento)
        db.session.commit()
        log_auditoria('Orçamento excluído', f'{orcamento.categoria} - {orcamento.mes}/{orcamento.ano}')
        flash('Orçamento excluído!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir orçamento: {str(e)}', 'danger')
    return redirect(url_for('financeiro.orcamento'))


@financeiro_bp.route('/financeiro/exportar/pdf')
@login_required
def financeiro_exportar_pdf():
    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    if not filtro_data_ini:
        filtro_data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not filtro_data_fim:
        filtro_data_fim = date.today().strftime('%Y-%m-%d')
    data_ini = datetime.strptime(filtro_data_ini, '%Y-%m-%d').date()
    data_fim = datetime.strptime(filtro_data_fim, '%Y-%m-%d').date()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini, Despesa.data <= data_fim
    ).order_by(Despesa.data.desc()).all()
    total = sum(d.valor for d in despesas)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=20, textColor=colors.HexColor('#27AE60'),
                                     spaceAfter=20, alignment=TA_CENTER)
        elements.append(Paragraph('Terra Roxa System', title_style))
        elements.append(Paragraph(f'Relatório de Despesas - {filtro_data_ini} a {filtro_data_fim}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(f'<b>Total de Registros:</b> {len(despesas)}  |  <b>Total:</b> R$ {float(total):.2f}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        data_table = [['Data', 'Descrição', 'Categoria', 'Valor']]
        for d in despesas:
            data_table.append([
                d.data.strftime('%d/%m/%Y'),
                d.descricao,
                d.categoria or '-',
                f'R$ {float(d.valor):.2f}'
            ])
        t = Table(data_table, colWidths=[3*cm, 6*cm, 3*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F39C12')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.Color(0,0,0, alpha=0.3)),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF9E7')])
        ]))
        elements.append(t)
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} - Terra Roxa System', styles['Normal']))
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Disposition'] = f'attachment; filename=despesas_{filtro_data_ini}_{filtro_data_fim}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response
    except Exception:
        html = f'''
        <html><head><title>Despesas</title><style>
        body{{font-family:Arial;padding:40px}} h1{{color:#F39C12}}
        table{{border-collapse:collapse;width:100%;margin:20px 0}}
        th,td{{border:1px solid #ddd;padding:12px;text-align:center}} th{{background:#F39C12;color:#fff}}
        </style></head><body>
        <h1>Terra Roxa System</h1><h3>Relatório de Despesas</h3>
        <p><strong>Período:</strong> {filtro_data_ini} a {filtro_data_fim}</p>
        <p>Total: {len(despesas)} registros | R$ {float(total):.2f}</p><hr>
        <table><tr><th>Data</th><th>Descrição</th><th>Categoria</th><th>Valor</th></tr>'''
        for d in despesas:
            html += f'<tr><td>{d.data.strftime("%d/%m/%Y")}</td><td>{d.descricao}</td><td>{d.categoria or "-"}</td><td>R$ {float(d.valor):.2f}</td></tr>'
        html += '</table></body></html>'
        response = make_response(html)
        response.headers['Content-Disposition'] = f'attachment; filename=despesas_{filtro_data_ini}_{filtro_data_fim}.html'
        response.headers['Content-Type'] = 'text/html'
        return response


@financeiro_bp.route('/financeiro/exportar/excel')
@login_required
def financeiro_exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    filtro_data_ini = request.args.get('data_ini')
    filtro_data_fim = request.args.get('data_fim')
    if not filtro_data_ini:
        filtro_data_ini = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not filtro_data_fim:
        filtro_data_fim = date.today().strftime('%Y-%m-%d')
    data_ini = datetime.strptime(filtro_data_ini, '%Y-%m-%d').date()
    data_fim = datetime.strptime(filtro_data_fim, '%Y-%m-%d').date()
    despesas = Despesa.query.filter(
        Despesa.data >= data_ini, Despesa.data <= data_fim
    ).order_by(Despesa.data.desc()).all()
    total = sum(d.valor for d in despesas)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Despesas'
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(1, 1, 'Terra Roxa System - Relatorio de Despesas').font = Font(bold=True, size=14, color='F39C12')
    ws.merge_cells('A1:D1')
    ws.cell(2, 1, f'Periodo: {filtro_data_ini} a {filtro_data_fim}')
    ws.merge_cells('A2:D2')
    headers = ['Data', 'Descricao', 'Categoria', 'Valor']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for i, d in enumerate(despesas, 5):
        ws.cell(i, 1, d.data.strftime('%d/%m/%Y')).border = thin_border
        ws.cell(i, 2, d.descricao).border = thin_border
        ws.cell(i, 3, d.categoria or '-').border = thin_border
        ws.cell(i, 4, float(d.valor)).border = thin_border
    row = 5 + len(despesas)
    ws.cell(row, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(row, 1).border = thin_border
    ws.cell(row, 4, float(total)).font = Font(bold=True)
    ws.cell(row, 4).border = thin_border
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 20
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = f'attachment; filename=despesas_{filtro_data_ini}_{filtro_data_fim}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
